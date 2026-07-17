from __future__ import annotations

import copy
import hashlib
import json
import os
import re
import time
import urllib.error
import urllib.request
import zipfile
import xml.etree.ElementTree as ET
import uuid
from contextlib import contextmanager
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROJECT_FOLDER = "AI_PMO_TASKS"
LEGACY_PROJECT_FOLDERS: tuple[str, ...] = ()
MOSCOW = timezone(timedelta(hours=3))
ACTIVE_STATUSES = {
    "Registered",
    "In_Progress",
    "In Progress",
    "Result_Submitted",
    "Validating",
    "Rework_Required",
    "Rework Required",
    "Human_Review_Required",
    "Human Review Required",
}
SUPPORTED_TEXT_SUFFIXES = {".txt", ".md", ".docx", ".pdf"}
RESULT_FILE_RE = re.compile(
    r"(?P<task_id>TASK-\d{8}-\d{3})_v(?P<version>\d+)",
    re.IGNORECASE,
)


def now_moscow() -> datetime:
    return datetime.now(MOSCOW)


def iso(value: datetime | date | None = None) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=MOSCOW)
        return value.isoformat()
    return value.isoformat()


def parse_dt(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=value.tzinfo or MOSCOW)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), MOSCOW)
    if not value:
        return None
    text = str(value).strip()
    for candidate in (text, text.replace("Z", "+00:00")):
        try:
            parsed = datetime.fromisoformat(candidate)
            return parsed.replace(tzinfo=parsed.tzinfo or MOSCOW)
        except Exception:
            pass
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=MOSCOW)
        except Exception:
            pass
    return None


def parse_date(value: Any) -> date | None:
    parsed = parse_dt(value)
    return parsed.date() if parsed else None


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def read_document(path: Path) -> str:
    """Read supported source files without OCR."""
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return read_text(path)

    if suffix == ".docx":
        try:
            with zipfile.ZipFile(path) as archive:
                xml = archive.read("word/document.xml")
            root = ET.fromstring(xml)
            namespace = {
                "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
            }
            paragraphs = []
            for paragraph in root.findall(".//w:p", namespace):
                parts = [
                    node.text or ""
                    for node in paragraph.findall(".//w:t", namespace)
                ]
                text = "".join(parts).strip()
                if text:
                    paragraphs.append(text)
            return "\n".join(paragraphs)
        except Exception as exc:
            raise RuntimeError(
                f"Не удалось прочитать DOCX {path.name}: {exc}"
            ) from exc

    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
        except Exception as exc:
            raise RuntimeError(
                "Для чтения PDF не найден пакет pypdf. "
                "Используйте TXT, MD или DOCX."
            ) from exc
        try:
            reader = PdfReader(str(path))
            text = "\n".join(
                (page.extract_text() or "").strip()
                for page in reader.pages
            ).strip()
            if not text:
                raise RuntimeError(
                    "PDF не содержит извлекаемого текстового слоя."
                )
            return text
        except Exception as exc:
            raise RuntimeError(
                f"Не удалось прочитать PDF {path.name}: {exc}"
            ) from exc

    raise RuntimeError(f"Неподдерживаемый формат: {path.suffix}")


def normalize_text(value: Any) -> str:
    text = str(value or "").lower().replace("ё", "е")
    text = re.sub(r"[^a-zа-я0-9\-]+", " ", text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip()



class LLMServiceError(RuntimeError):
    pass


def _host_service_request(
    route: str,
    payload: dict[str, Any],
    *,
    timeout: int,
) -> dict[str, Any]:
    base_url = str(os.environ.get("HOST_SERVICE_URL") or "").strip()
    token = str(os.environ.get("HOST_SERVICE_TOKEN") or "").strip()

    if not base_url or not token:
        raise LLMServiceError(
            "Host Service недоступен. Проверьте разрешение inject_chat "
            "у task-control-dashboard и перезапустите Ouroboros."
        )

    request = urllib.request.Request(
        f"{base_url.rstrip('/')}{route}",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        method="POST",
        headers={
            "Content-Type": "application/json",
            "x-skill-token": token,
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8", errors="replace")
            data = json.loads(raw or "{}")
            if not isinstance(data, dict):
                raise LLMServiceError(
                    "Host Service вернул ответ неизвестного формата."
                )
            return data
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise LLMServiceError(
            f"Host Service HTTP {exc.code}: {details[:800]}"
        ) from exc
    except urllib.error.URLError as exc:
        raise LLMServiceError(
            f"Host Service недоступен: {exc.reason}"
        ) from exc


def _fresh_internal_chat(purpose: str) -> int:
    response = _host_service_request(
        "/chat/allocate-internal",
        {
            # Host Service accepts only predefined internal-chat ranges.
            # "a2a" is the supported general-purpose isolated range.
            "range_name": "a2a"
        },
        timeout=20,
    )
    chat_id = int(response.get("chat_id") or 0)
    if not chat_id:
        raise LLMServiceError(
            "Host Service не выделил новый internal chat_id."
        )
    return chat_id


def _extract_json_object(raw: str) -> dict[str, Any]:
    text = str(raw or "").strip()
    if not text:
        raise LLMServiceError("Языковая модель вернула пустой ответ.")

    fence = re.search(
        r"```(?:json)?\s*(\{.*\})\s*```",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if fence:
        text = fence.group(1).strip()

    try:
        value = json.loads(text)
        if isinstance(value, dict):
            return value
    except Exception:
        pass

    start = text.find("{")
    if start >= 0:
        depth = 0
        in_string = False
        escaped = False
        for index in range(start, len(text)):
            char = text[index]
            if in_string:
                if escaped:
                    escaped = False
                elif char == "\\":
                    escaped = True
                elif char == '"':
                    in_string = False
                continue
            if char == '"':
                in_string = True
            elif char == "{":
                depth += 1
            elif char == "}":
                depth -= 1
                if depth == 0:
                    candidate = text[start:index + 1]
                    try:
                        value = json.loads(candidate)
                        if isinstance(value, dict):
                            return value
                    except Exception:
                        break

    raise LLMServiceError(
        "Языковая модель не вернула валидный JSON. "
        f"Начало ответа: {text[:700]}"
    )


def llm_json(
    *,
    purpose: str,
    prompt: str,
    timeout_seconds: int = 600,
) -> dict[str, Any]:
    """Run one semantic operation in a fresh, isolated LLM context."""
    chat_id = _fresh_internal_chat(purpose)
    response = _host_service_request(
        "/chat/inject",
        {
            "chat_id": chat_id,
            "user_id": 0,
            "sender_label": "AI PMO LLM Runtime",
            "text": prompt,
            "wait_for_response": True,
            "timeout_sec": timeout_seconds,
        },
        timeout=timeout_seconds + 30,
    )
    if not response.get("ok"):
        raise LLMServiceError(
            str(response.get("error") or "LLM-вызов не выполнен.")
        )
    return _extract_json_object(str(response.get("response") or ""))


def _bounded_text(text: str, limit: int = 60000) -> str:
    value = str(text or "")
    if len(value) <= limit:
        return value
    return value[:limit] + "\n\n[Текст обрезан локальным orchestrator]"


def _date_iso(value: Any) -> date | None:
    if value in (None, "", "null"):
        return None
    text = str(value).strip()
    try:
        return date.fromisoformat(text)
    except Exception:
        return parse_date(text)


def _task_sequence(existing_ids: set[str], task_date: date) -> int:
    prefix = f"TASK-{task_date:%Y%m%d}-"
    values = []
    for task_id in existing_ids:
        if task_id.startswith(prefix):
            try:
                values.append(int(task_id.rsplit("-", 1)[-1]))
            except Exception:
                pass
    return max(values, default=0) + 1


def _people_context(store: "WorkbookStore") -> list[dict[str, str]]:
    book = store.open("02_people.xlsx")
    people = [
        {
            "person_id": str(row.get("person_id") or ""),
            "full_name": str(row.get("full_name") or ""),
            "email": str(row.get("email") or ""),
        }
        for _, row in iter_rows(book["People"])
    ]
    aliases: dict[str, list[str]] = {}
    for _, row in iter_rows(book["Aliases"]):
        person_id = str(row.get("person_id") or "")
        alias = str(row.get("alias_text") or "")
        if person_id and alias:
            aliases.setdefault(person_id, []).append(alias)
    for person in people:
        person["aliases"] = "; ".join(
            aliases.get(person["person_id"], [])
        )
    return people


def _projects_context(store: "WorkbookStore") -> list[dict[str, str]]:
    book = store.open("01_projects.xlsx")
    return [
        {
            "project_id": str(row.get("project_id") or ""),
            "project_name": str(row.get("project_name") or ""),
            "priority": str(row.get("priority") or ""),
        }
        for _, row in iter_rows(book["Projects"])
    ]


def _is_workspace(candidate: Path) -> bool:
    return (candidate / "registry").is_dir()


def _windows_documents_dir() -> Path | None:
    """Return the current user's real Windows Documents known folder.

    This supports redirected folders and OneDrive without embedding a username.
    """
    if os.name != "nt":
        return None
    try:
        import ctypes
        from ctypes import wintypes
        from uuid import UUID

        # FOLDERID_Documents = {FDD39AD0-238F-46AF-ADB4-6C85480369C7}
        raw = UUID("fdd39ad0-238f-46af-adb4-6c85480369c7").bytes_le
        guid = (ctypes.c_ubyte * 16).from_buffer_copy(raw)
        path_ptr = ctypes.c_wchar_p()
        shell32 = ctypes.WinDLL("shell32", use_last_error=True)
        ole32 = ctypes.WinDLL("ole32", use_last_error=True)
        shell32.SHGetKnownFolderPath.argtypes = [
            ctypes.POINTER(ctypes.c_ubyte * 16),
            wintypes.DWORD,
            wintypes.HANDLE,
            ctypes.POINTER(ctypes.c_wchar_p),
        ]
        shell32.SHGetKnownFolderPath.restype = ctypes.HRESULT
        result = shell32.SHGetKnownFolderPath(
            ctypes.byref(guid), 0, None, ctypes.byref(path_ptr)
        )
        if result != 0 or not path_ptr.value:
            return None
        try:
            return Path(path_ptr.value)
        finally:
            ole32.CoTaskMemFree(path_ptr)
    except Exception:
        return None


def _user_documents_candidates() -> list[Path]:
    """Return portable user Documents candidates in priority order.

    Windows uses the Known Folder API (including redirected/OneDrive folders).
    macOS and Linux use the current home directory. On macOS an iCloud Drive
    Documents candidate is also recognised when it exists.
    """
    candidates: list[Path] = []

    known = _windows_documents_dir()
    if known is not None:
        candidates.append(known)

    # Cross-platform fallbacks. No concrete username or drive is embedded.
    userprofile = str(os.environ.get("USERPROFILE") or "").strip()
    if userprofile:
        candidates.append(Path(userprofile).expanduser() / "Documents")

    home = Path.home()
    candidates.append(home / "Documents")

    # Optional macOS iCloud Drive location. This does not override ~/Documents;
    # it is only a fallback for users who keep Documents in iCloud Drive.
    if sys.platform == "darwin":
        candidates.append(
            home
            / "Library"
            / "Mobile Documents"
            / "com~apple~CloudDocs"
            / "Documents"
        )

    unique: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        try:
            resolved = candidate.expanduser().resolve()
        except OSError:
            resolved = candidate.expanduser().absolute()
        key = os.path.normcase(str(resolved))
        if key not in seen:
            seen.add(key)
            unique.append(resolved)
    return unique


def _candidate_data_roots() -> list[Path]:
    """Return fallback Ouroboros data roots for backward compatibility."""
    candidates: list[Path] = []

    configured_data = str(os.environ.get("OUROBOROS_DATA_DIR") or "").strip()
    if configured_data:
        candidates.append(Path(configured_data).expanduser())

    configured_root = str(os.environ.get("OUROBOROS_ROOT") or "").strip()
    if configured_root:
        candidates.append(Path(configured_root).expanduser() / "data")

    here = Path(__file__).resolve()
    for parent in here.parents:
        if parent.name.lower() == "data":
            candidates.append(parent)
            break
        if (parent / "skills" / "external").is_dir():
            candidates.append(parent)

    unique: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except OSError:
            resolved = candidate.absolute()
        key = os.path.normcase(str(resolved))
        if key not in seen:
            seen.add(key)
            unique.append(resolved)
    return unique


def find_project_root() -> Path:
    """Resolve AI_PMO_TASKS without machine-specific absolute paths.

    Priority:
    1. AI_PMO_TASKS_ROOT explicit override;
    2. the current user's real system Documents/AI_PMO_TASKS folder;
    3. Ouroboros data/documents/AI_PMO_TASKS for backward compatibility.
    """
    configured = str(os.environ.get("AI_PMO_TASKS_ROOT") or "").strip()
    if not configured:
        configured = str(os.environ.get("AI_PMO_TASKS_DOCUMENTS") or "").strip()
    if configured:
        candidate = Path(configured).expanduser()
        if _is_workspace(candidate):
            return candidate.resolve()
        raise FileNotFoundError(
            "AI_PMO_TASKS_ROOT задан, но в указанной папке нет registry: "
            f"{candidate}"
        )

    checked: list[Path] = []

    # Preferred final layout: <system Documents>/AI_PMO_TASKS.
    for documents_dir in _user_documents_candidates():
        candidate = documents_dir / PROJECT_FOLDER
        checked.append(candidate)
        if _is_workspace(candidate):
            return candidate.resolve()

    # Backward-compatible layout used by earlier packages.
    for data_root in _candidate_data_roots():
        candidate = data_root / "documents" / PROJECT_FOLDER
        checked.append(candidate)
        if _is_workspace(candidate):
            return candidate.resolve()

    locations = "\n - ".join(str(path) for path in checked) or "нет кандидатов"
    raise FileNotFoundError(
        "Не найдена рабочая папка AI_PMO_TASKS. Проверены:\n - "
        + locations
        + "\nСкопируйте AI_PMO_TASKS в папку Documents текущего пользователя "
          "или задайте переменную AI_PMO_TASKS_ROOT."
    )

def ensure_structure(root: Path) -> None:
    required_dirs = (
        root / "registry",
        root / "input" / "meetings",
        root / "input" / "results",
        root / "input" / "commands",
        root / "archive" / "commands" / "processed",
        root / "archive" / "commands" / "failed",
        root / "archive" / "meetings",
        root / "archive" / "results",
        root / "output" / "dashboards",
        root / "output" / "reports",
    )
    for directory in required_dirs:
        directory.mkdir(parents=True, exist_ok=True)

    required_files = (
        "01_projects.xlsx",
        "02_people.xlsx",
        "03_task_register.xlsx",
        "04_task_rules.xlsx",
        "05_message_outbox.xlsx",
        "06_task_event_queue.xlsx",
        "07_task_control_state.xlsx",
        "08_skill_action_log.xlsx",
        "09_source_register.xlsx",
        "10_loop_trigger_settings.xlsx",
    )
    missing = [
        name for name in required_files
        if not (root / "registry" / name).is_file()
    ]
    if missing:
        raise FileNotFoundError(
            "Не найдены обязательные Excel-реестры: " + ", ".join(missing)
        )


@contextmanager
def processor_lock(root: Path, stale_after_seconds: int = 1800):
    lock_path = root / "archive" / ".processor.lock"
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "pid": os.getpid(),
        "created_at": iso(now_moscow()),
    }

    while True:
        try:
            descriptor = os.open(
                str(lock_path),
                os.O_CREAT | os.O_EXCL | os.O_WRONLY,
            )
            with os.fdopen(descriptor, "w", encoding="utf-8") as stream:
                json.dump(payload, stream, ensure_ascii=False)
            break
        except FileExistsError:
            try:
                age = time.time() - lock_path.stat().st_mtime
            except OSError:
                age = 0
            if age > stale_after_seconds:
                try:
                    lock_path.unlink()
                    continue
                except OSError:
                    pass
            raise RuntimeError(
                "Другой запуск уже обрабатывает поручения. "
                "Повторите через несколько секунд."
            )

    try:
        yield
    finally:
        try:
            lock_path.unlink()
        except OSError:
            pass


class WorkbookStore:
    def __init__(self, root: Path):
        self.root = root
        self.books: dict[str, Any] = {}
        self.paths: dict[str, Path] = {}
        self.dirty: set[str] = set()

    def open(self, filename: str):
        if filename not in self.books:
            path = self.root / "registry" / filename
            self.books[filename] = load_workbook(path)
            self.paths[filename] = path
        return self.books[filename]

    def mark_dirty(self, filename: str) -> None:
        self.dirty.add(filename)

    def save_all(self) -> None:
        for filename in sorted(self.dirty):
            workbook = self.books[filename]
            path = self.paths[filename]
            temp = path.with_name(
                f".{path.stem}.{uuid.uuid4().hex[:8]}.tmp.xlsx"
            )
            workbook.save(temp)
            os.replace(temp, path)
        self.dirty.clear()

    def close(self) -> None:
        for workbook in self.books.values():
            try:
                workbook.close()
            except Exception:
                pass


def headers(ws) -> list[str]:
    return [str(cell.value or "").strip() for cell in ws[3]]


def iter_rows(ws) -> Iterable[tuple[int, dict[str, Any]]]:
    names = headers(ws)
    for row_index in range(4, ws.max_row + 1):
        values = [ws.cell(row_index, col + 1).value for col in range(len(names))]
        if not any(value not in (None, "") for value in values):
            continue
        yield row_index, {
            name: values[index]
            for index, name in enumerate(names)
            if name
        }


def find_row(ws, key: str, value: Any) -> tuple[int, dict[str, Any]] | None:
    target = str(value or "")
    for row_index, row in iter_rows(ws):
        if str(row.get(key) or "") == target:
            return row_index, row
    return None


def append_mapping(ws, mapping: dict[str, Any]) -> int:
    names = headers(ws)
    row_index = max(ws.max_row + 1, 4)
    for column, name in enumerate(names, start=1):
        ws.cell(row_index, column).value = mapping.get(name)
    return row_index


def update_mapping(ws, row_index: int, mapping: dict[str, Any]) -> None:
    names = headers(ws)
    positions = {name: index + 1 for index, name in enumerate(names) if name}
    for name, value in mapping.items():
        column = positions.get(name)
        if column:
            ws.cell(row_index, column).value = value


def load_settings(store: WorkbookStore) -> dict[str, Any]:
    settings = {
        "active_profile": "HYBRID",
        "widget_commands_enabled": True,
        "file_events_enabled": True,
        "periodic_full_run_enabled": True,
        "watcher_poll_seconds": 2,
        "periodic_interval_minutes": 180,
        "debounce_seconds": 15,
        "max_items_per_run": 100,
        "timezone": "Europe/Moscow",
        "archive_processed_source_files": False,
    }
    workbook = store.open("10_loop_trigger_settings.xlsx")
    ws = workbook["Control_Panel"]
    for _, row in iter_rows(ws):
        key = str(row.get("setting_name") or "").strip()
        if not key:
            continue
        value = row.get("setting_value")
        value_type = str(row.get("value_type") or "Text")
        if value_type == "Boolean":
            settings[key] = str(value).strip().lower() in {
                "true", "1", "yes", "да"
            }
        elif value_type == "Integer":
            try:
                settings[key] = int(float(value))
            except Exception:
                pass
        else:
            settings[key] = str(value or "").strip()
    settings["active_profile"] = str(
        settings.get("active_profile") or "HYBRID"
    ).upper()
    return settings


def project_lookup(store: WorkbookStore) -> tuple[dict[str, dict], dict[str, dict]]:
    workbook = store.open("01_projects.xlsx")
    rows = [row for _, row in iter_rows(workbook["Projects"])]
    by_id = {str(row.get("project_id")): row for row in rows if row.get("project_id")}
    by_name = {
        normalize_text(row.get("project_name")): row
        for row in rows if row.get("project_name")
    }
    return by_id, by_name


def people_lookup(store: WorkbookStore) -> tuple[dict[str, dict], dict[str, dict]]:
    workbook = store.open("02_people.xlsx")
    people = [row for _, row in iter_rows(workbook["People"])]
    by_id = {str(row.get("person_id")): row for row in people if row.get("person_id")}
    by_name: dict[str, dict] = {
        normalize_text(row.get("full_name")): row
        for row in people if row.get("full_name")
    }
    for _, alias in iter_rows(workbook["Aliases"]):
        person = by_id.get(str(alias.get("person_id") or ""))
        if person and alias.get("alias_text"):
            by_name[normalize_text(alias.get("alias_text"))] = person
    return by_id, by_name


def extract_meeting_metadata(text: str) -> dict[str, Any]:
    meeting_date = None
    match = re.search(
        r"Дата\s+встречи\s*:\s*(\d{2}\.\d{2}\.\d{4})",
        text,
        flags=re.IGNORECASE,
    )
    if match:
        meeting_date = parse_date(match.group(1))

    project_id = None
    match = re.search(
        r"Project\s*ID\s*:\s*([A-Za-z0-9_-]+)",
        text,
        flags=re.IGNORECASE,
    )
    if match:
        project_id = match.group(1).strip()

    project_name = None
    match = re.search(
        r"^\s*Проект\s*:\s*(.+?)\s*$",
        text,
        flags=re.IGNORECASE | re.MULTILINE,
    )
    if match:
        project_name = match.group(1).strip()

    title = ""
    for line in text.splitlines():
        if line.strip():
            title = line.strip()
            break

    participants = []
    participant_block = re.search(
        r"Участники\s*:\s*(.*?)(?=^\s*\d+\.\s|\Z)",
        text,
        flags=re.IGNORECASE | re.MULTILINE | re.DOTALL,
    )
    if participant_block:
        participants = [
            re.sub(r"^\s*[-•]\s*", "", line).strip()
            for line in participant_block.group(1).splitlines()
            if line.strip()
        ]

    return {
        "meeting_date": meeting_date,
        "project_id": project_id,
        "project_name": project_name,
        "meeting_title": title,
        "participants": participants,
    }


def extract_assignments(text: str) -> list[dict[str, Any]]:
    assignments: list[dict[str, Any]] = []
    blocks = re.findall(
        r"(?ms)^\s*(\d+)\.\s*(.*?)(?=^\s*\d+\.\s|\Z)",
        text,
    )
    order = 0
    for source_order_text, block in blocks:
        lines = [line.rstrip() for line in block.strip().splitlines()]
        if not lines:
            continue
        first_line = lines[0].strip()
        match = re.match(
            r"^(?P<assignee>.+?)\s+до\s+"
            r"(?P<due>\d{2}\.\d{2}\.\d{4})\s+"
            r"(?P<description>.+?)\.?$",
            first_line,
            flags=re.IGNORECASE,
        )
        if not match:
            continue

        order += 1
        expected_result = ""
        expected_match = re.search(
            r"Ожидаемый\s+результат\s*:\s*(.+?)(?="
            r"\n\s*Критерии\s+при[её]мки\s*:|\Z)",
            block,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if expected_match:
            expected_result = " ".join(
                expected_match.group(1).strip().split()
            )

        criteria: list[str] = []
        criteria_match = re.search(
            r"Критерии\s+при[её]мки\s*:\s*(.*)$",
            block,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if criteria_match:
            for line in criteria_match.group(1).splitlines():
                criterion = re.sub(r"^\s*[-•]\s*", "", line).strip()
                if criterion:
                    criteria.append(criterion.rstrip(";."))

        description = match.group("description").strip().rstrip(".")
        assignments.append(
            {
                "source_order": int(source_order_text),
                "assignment_order": order,
                "assignee_raw": match.group("assignee").strip(),
                "due_date": parse_date(match.group("due")),
                "due_date_text": match.group("due"),
                "task_description": description,
                "task_title": description[:120],
                "expected_result": expected_result,
                "criteria": criteria,
                "source_fragment": block.strip(),
            }
        )
    return assignments


def timing_state(due_date: date | None, status: str, today: date) -> str:
    if status in {"Closed", "Cancelled"}:
        return "Closed"
    if due_date is None:
        return "Unknown"
    delta = (due_date - today).days
    if delta < 0:
        return "Overdue"
    if delta == 0:
        return "Due_Today"
    if delta <= 3:
        return "Due_Soon"
    return "On_Track"


def append_action_log(
    store: WorkbookStore,
    *,
    run_id: str,
    skill_id: str,
    action_type: str,
    object_type: str,
    object_id: str,
    project_id: str = "",
    status: str = "Success",
    source_file: str = "",
    target_file: str = "",
    target_sheet: str = "",
    details: str = "",
    error_code: str = "",
) -> None:
    workbook = store.open("08_skill_action_log.xlsx")
    ws = workbook["Action_Log"]
    timestamp = now_moscow()
    append_mapping(
        ws,
        {
            "log_id": f"LOG-{timestamp:%Y%m%d%H%M%S}-{uuid.uuid4().hex[:6]}",
            "logged_at": iso(timestamp),
            "run_id": run_id,
            "control_id": "task_lifecycle_pmo",
            "skill_id": skill_id,
            "action_type": action_type,
            "object_type": object_type,
            "object_id": object_id,
            "project_id": project_id,
            "action_status": status,
            "source_file": source_file,
            "source_sheet": "",
            "target_file": target_file,
            "target_sheet": target_sheet,
            "before_value": "",
            "after_value": "",
            "details": details,
            "rule_id": "HYBRID_LLM_RUNTIME_V4",
            "idempotency_key": hashlib.sha256(
                f"{run_id}|{skill_id}|{action_type}|{object_id}".encode("utf-8")
            ).hexdigest(),
            "error_code": error_code,
        },
    )
    store.mark_dirty("08_skill_action_log.xlsx")


def append_history(
    store: WorkbookStore,
    *,
    task_id: str,
    run_id: str,
    from_status: str,
    to_status: str,
    reason: str,
    skill_id: str,
    details: str = "",
) -> None:
    workbook = store.open("03_task_register.xlsx")
    ws = workbook["Task_History"]
    timestamp = now_moscow()
    append_mapping(
        ws,
        {
            "history_id": f"HIS-{timestamp:%Y%m%d%H%M%S}-{uuid.uuid4().hex[:6]}",
            "task_id": task_id,
            "changed_at": iso(timestamp),
            "run_id": run_id,
            "from_status": from_status,
            "to_status": to_status,
            "change_reason": reason,
            "skill_id": skill_id,
            "details": details,
        },
    )
    store.mark_dirty("03_task_register.xlsx")


def process_meetings(
    store: WorkbookStore,
    root: Path,
    run_id: str,
    limit: int,
) -> dict[str, int]:
    metrics = {
        "sources_processed": 0,
        "assignments_found": 0,
        "tasks_registered": 0,
    }
    by_project_id, by_project_name = project_lookup(store)
    by_person_id, by_person_name = people_lookup(store)
    projects_context = _projects_context(store)
    people_context = _people_context(store)

    source_book = store.open("09_source_register.xlsx")
    source_ws = source_book["Source_Documents"]
    candidate_ws = source_book["Extracted_Candidates"]
    known_hashes = {
        str(row.get("file_hash") or "").lower()
        for _, row in iter_rows(source_ws)
        if row.get("file_hash")
    }

    task_book = store.open("03_task_register.xlsx")
    tasks_ws = task_book["Tasks"]
    criteria_ws = task_book["Acceptance_Criteria"]
    existing_task_ids = {
        str(row.get("task_id"))
        for _, row in iter_rows(tasks_ws)
        if row.get("task_id")
    }

    files = [
        path
        for path in sorted((root / "input" / "meetings").iterdir())
        if path.is_file()
        and not path.name.lower().startswith("readme")
        and path.suffix.lower() in SUPPORTED_TEXT_SUFFIXES
    ][:limit]

    for path in files:
        file_hash = sha256_file(path)
        if file_hash.lower() in known_hashes:
            continue

        protocol_text = read_document(path)
        local_metadata = extract_meeting_metadata(protocol_text)
        current_date = now_moscow().date()

        prompt = f"""
MEETING_EXTRACTION_V4

Ты — AI-аналитик проектного офиса. Проанализируй протокол встречи и верни
ТОЛЬКО один JSON-объект без Markdown и комментариев.

Нужно:
1. Отличить поручения от решений, информации и открытых вопросов.
2. Извлечь каждое реальное поручение, даже если оно сформулировано свободно.
3. Разрешить относительные даты относительно даты встречи.
4. Сопоставить проект и ответственного со справочниками, если это возможно.
5. Не придумывать поручения.
6. Если критерии приёмки не указаны, сформировать 1–4 измеримых критерия на
   основании ожидаемого результата.
7. Сохранить короткий исходный фрагмент для аудита.

Текущая дата: {current_date.isoformat()}
Локально найденная дата встречи:
{iso(local_metadata.get("meeting_date")) or "не определена"}

Справочник проектов:
{json.dumps(projects_context, ensure_ascii=False)}

Справочник людей:
{json.dumps(people_context, ensure_ascii=False)}

JSON-схема:
{{
  "meeting_date": "YYYY-MM-DD или null",
  "project_id": "ID из справочника или пустая строка",
  "project_name": "название проекта",
  "meeting_title": "краткое название встречи",
  "participants": ["ФИО"],
  "assignments": [
    {{
      "source_order": 1,
      "task_title": "краткий заголовок",
      "task_description": "конкретное действие",
      "assignee_raw": "как указано в протоколе",
      "assignee_id": "ID из справочника или пустая строка",
      "due_date": "YYYY-MM-DD или null",
      "expected_result": "ожидаемый артефакт или результат",
      "criteria": ["измеримый критерий"],
      "criteria_source": "Protocol или Generated",
      "source_fragment": "короткая цитата или точный пересказ фрагмента",
      "confidence": 0.0
    }}
  ]
}}

ПРОТОКОЛ:
---BEGIN---
{_bounded_text(protocol_text)}
---END---
""".strip()

        extracted = llm_json(
            purpose="meeting",
            prompt=prompt,
        )
        assignments = extracted.get("assignments")
        if not isinstance(assignments, list):
            raise LLMServiceError(
                f"LLM не вернула массив assignments для {path.name}."
            )

        meeting_date = (
            _date_iso(extracted.get("meeting_date"))
            or local_metadata.get("meeting_date")
            or current_date
        )
        project_id_from_llm = str(
            extracted.get("project_id") or ""
        ).strip()
        project_name_from_llm = str(
            extracted.get("project_name")
            or local_metadata.get("project_name")
            or ""
        ).strip()

        project = by_project_id.get(project_id_from_llm)
        if project is None and project_name_from_llm:
            project = by_project_name.get(
                normalize_text(project_name_from_llm)
            )
        project_id = str(
            (project or {}).get("project_id")
            or project_id_from_llm
            or local_metadata.get("project_id")
            or ""
        )
        project_priority = str(
            (project or {}).get("priority") or "High"
        )
        source_id = f"SRC-{meeting_date:%Y%m%d}-{file_hash[:6].upper()}"

        valid_assignments = []
        for raw in assignments:
            if not isinstance(raw, dict):
                continue
            description = str(
                raw.get("task_description") or ""
            ).strip()
            assignee_raw = str(raw.get("assignee_raw") or "").strip()
            if not description or not assignee_raw:
                continue
            valid_assignments.append(raw)

        append_mapping(
            source_ws,
            {
                "source_id": source_id,
                "source_type": "Meeting_Protocol",
                "file_name": path.name,
                "file_path": str(path.resolve()),
                "file_hash": file_hash,
                "meeting_date": meeting_date,
                "project_id": project_id,
                "meeting_title": str(
                    extracted.get("meeting_title")
                    or local_metadata.get("meeting_title")
                    or path.stem
                ),
                "participants": "\n".join(
                    str(item)
                    for item in (
                        extracted.get("participants")
                        if isinstance(extracted.get("participants"), list)
                        else local_metadata.get("participants") or []
                    )
                ),
                "status": "Processed",
                "processed_at": iso(now_moscow()),
                "run_id": run_id,
                "idempotency_key": file_hash,
                "error": "",
            },
        )
        store.mark_dirty("09_source_register.xlsx")
        known_hashes.add(file_hash.lower())
        metrics["sources_processed"] += 1
        metrics["assignments_found"] += len(valid_assignments)

        append_action_log(
            store,
            run_id=run_id,
            skill_id="SK-18",
            action_type="llm_parse_meeting",
            object_type="source",
            object_id=source_id,
            project_id=project_id,
            status="Success",
            source_file=path.name,
            target_file="09_source_register.xlsx",
            target_sheet="Source_Documents",
            details=f"LLM выделила поручений: {len(valid_assignments)}",
        )

        next_sequence = _task_sequence(
            existing_task_ids,
            meeting_date,
        )

        for assignment_order, assignment in enumerate(
            valid_assignments,
            start=1,
        ):
            source_order = int(
                assignment.get("source_order") or assignment_order
            )
            task_id = (
                f"TASK-{meeting_date:%Y%m%d}-{next_sequence:03d}"
            )
            next_sequence += 1

            candidate_id = (
                f"CAND-{meeting_date:%Y%m%d}-"
                f"{task_id.rsplit('-', 1)[-1]}-{file_hash[:4].upper()}"
            )

            assignee_id = str(
                assignment.get("assignee_id") or ""
            ).strip()
            person = by_person_id.get(assignee_id)
            if person is None:
                person = by_person_name.get(
                    normalize_text(assignment.get("assignee_raw"))
                )
            assignee_id = str((person or {}).get("person_id") or assignee_id)
            assignee_name = str(
                (person or {}).get("full_name")
                or assignment.get("assignee_raw")
                or ""
            )
            assignee_email = str((person or {}).get("email") or "")
            due_date = _date_iso(assignment.get("due_date"))
            criteria = [
                str(item).strip()
                for item in (
                    assignment.get("criteria")
                    if isinstance(assignment.get("criteria"), list)
                    else []
                )
                if str(item).strip()
            ]
            expected_result = str(
                assignment.get("expected_result") or ""
            ).strip()
            if not criteria:
                criteria = [
                    expected_result
                    or "Предоставлен результат, соответствующий поручению"
                ]
            confidence = max(
                0.0,
                min(1.0, float(assignment.get("confidence") or 0.7)),
            )
            source_fragment = str(
                assignment.get("source_fragment")
                or assignment.get("task_description")
                or ""
            ).strip()
            idempotency_key = hashlib.sha256(
                f"{file_hash}|{source_order}|{source_fragment}".encode(
                    "utf-8"
                )
            ).hexdigest()

            append_mapping(
                candidate_ws,
                {
                    "candidate_id": candidate_id,
                    "source_id": source_id,
                    "source_order": source_order,
                    "assignment_order": assignment_order,
                    "candidate_type": "Assignment",
                    "project_id": project_id,
                    "task_title": str(
                        assignment.get("task_title")
                        or assignment.get("task_description")
                        or ""
                    )[:120],
                    "task_description": str(
                        assignment.get("task_description") or ""
                    ),
                    "assignee_raw": str(
                        assignment.get("assignee_raw") or ""
                    ),
                    "assignee_id": assignee_id,
                    "due_date_text": str(
                        assignment.get("due_date") or ""
                    ),
                    "due_date": due_date,
                    "expected_result": expected_result,
                    "acceptance_criteria_json": json.dumps(
                        criteria,
                        ensure_ascii=False,
                    ),
                    "source_fragment": source_fragment,
                    "confidence_assignment": confidence,
                    "confidence_assignee": 1.0 if person else confidence,
                    "confidence_due_date": (
                        confidence if due_date else 0.0
                    ),
                    "confidence_expected_result": (
                        confidence if expected_result else 0.5
                    ),
                    "confidence_overall": confidence,
                    "assumptions": "",
                    "candidate_status": "Registered",
                    "task_id": task_id,
                    "fragment_hash": hashlib.sha256(
                        source_fragment.encode("utf-8")
                    ).hexdigest(),
                    "idempotency_key": idempotency_key,
                    "created_at": iso(now_moscow()),
                    "updated_at": iso(now_moscow()),
                },
            )
            store.mark_dirty("09_source_register.xlsx")

            status = "Registered"
            append_mapping(
                tasks_ws,
                {
                    "task_id": task_id,
                    "project_id": project_id,
                    "source_id": source_id,
                    "candidate_id": candidate_id,
                    "assignment_order": assignment_order,
                    "task_title": str(
                        assignment.get("task_title")
                        or assignment.get("task_description")
                        or ""
                    )[:120],
                    "task_description": str(
                        assignment.get("task_description") or ""
                    ),
                    "assignee_id": assignee_id,
                    "assignee_name": assignee_name,
                    "assignee_email": assignee_email,
                    "assigned_at": meeting_date,
                    "due_date": due_date,
                    "expected_result": expected_result,
                    "priority": project_priority,
                    "status": status,
                    "timing_state": timing_state(
                        due_date, status, now_moscow().date()
                    ),
                    "acceptance_status": "Pending",
                    "current_iteration": 0,
                    "rework_due_date": None,
                    "latest_result_id": "",
                    "escalation_level": 0,
                    "last_reminder_at": None,
                    "last_status_at": iso(now_moscow()),
                    "closed_at": None,
                    "source_file": path.name,
                    "source_fragment": source_fragment,
                    "confidence_overall": confidence,
                    "idempotency_key": idempotency_key,
                    "created_at": iso(now_moscow()),
                    "updated_at": iso(now_moscow()),
                },
            )
            existing_task_ids.add(task_id)
            store.mark_dirty("03_task_register.xlsx")

            criteria_source = str(
                assignment.get("criteria_source") or "Generated"
            )
            if criteria_source not in {"Protocol", "Generated"}:
                criteria_source = "Generated"

            for criterion_order, criterion in enumerate(
                criteria,
                start=1,
            ):
                append_mapping(
                    criteria_ws,
                    {
                        "criterion_id": (
                            f"CRIT-{task_id}-{criterion_order:02d}"
                        ),
                        "task_id": task_id,
                        "criterion_order": criterion_order,
                        "criterion_text": criterion,
                        "mandatory": True,
                        "source_type": criteria_source,
                        "status": "Pending",
                        "evidence": "",
                        "result_id": "",
                        "checked_at": None,
                        "idempotency_key": hashlib.sha256(
                            (
                                f"{task_id}|{criterion_order}|"
                                f"{criterion}"
                            ).encode("utf-8")
                        ).hexdigest(),
                    },
                )
            store.mark_dirty("03_task_register.xlsx")

            append_history(
                store,
                task_id=task_id,
                run_id=run_id,
                from_status="",
                to_status="Registered",
                reason="Поручение извлечено языковой моделью",
                skill_id="SK-12",
                details=f"Источник: {path.name}",
            )
            append_action_log(
                store,
                run_id=run_id,
                skill_id="SK-12",
                action_type="register_llm_task",
                object_type="task",
                object_id=task_id,
                project_id=project_id,
                source_file=path.name,
                target_file="03_task_register.xlsx",
                target_sheet="Tasks",
                details=(
                    f"Ответственный: {assignee_name}; "
                    f"срок: {due_date}; confidence={confidence:.2f}"
                ),
            )
            metrics["tasks_registered"] += 1

    return metrics


def process_results(
    store: WorkbookStore,
    root: Path,
    run_id: str,
    limit: int,
) -> dict[str, int]:
    metrics = {
        "results_checked": 0,
        "accepted": 0,
        "rework_required": 0,
        "tasks_closed": 0,
        "human_review_required": 0,
    }

    task_book = store.open("03_task_register.xlsx")
    tasks_ws = task_book["Tasks"]
    criteria_ws = task_book["Acceptance_Criteria"]
    results_ws = task_book["Task_Results"]

    tasks = {
        str(row.get("task_id")): (row_index, row)
        for row_index, row in iter_rows(tasks_ws)
        if row.get("task_id")
    }

    source_book = store.open("09_source_register.xlsx")
    inbox_ws = source_book["Result_Inbox"]
    known_hashes = {
        str(row.get("file_hash") or "").lower()
        for _, row in iter_rows(inbox_ws)
        if row.get("file_hash")
    }

    files = [
        path
        for path in sorted((root / "input" / "results").iterdir())
        if path.is_file()
        and not path.name.lower().startswith("readme")
        and path.suffix.lower() in SUPPORTED_TEXT_SUFFIXES
    ][:limit]

    for path in files:
        match = RESULT_FILE_RE.search(path.stem)
        if not match:
            append_action_log(
                store,
                run_id=run_id,
                skill_id="SK-13",
                action_type="skip_result",
                object_type="file",
                object_id=path.name,
                status="Failed",
                source_file=path.name,
                details=(
                    "Имя результата должно иметь формат "
                    "TASK-YYYYMMDD-NNN_vN.ext"
                ),
                error_code="INVALID_RESULT_NAME",
            )
            continue

        task_id = match.group("task_id").upper()
        version = int(match.group("version"))
        if task_id not in tasks:
            append_action_log(
                store,
                run_id=run_id,
                skill_id="SK-13",
                action_type="defer_result",
                object_type="result",
                object_id=path.name,
                status="Skipped",
                source_file=path.name,
                details=f"Поручение {task_id} пока не зарегистрировано.",
                error_code="TASK_NOT_FOUND",
            )
            continue

        file_hash = sha256_file(path)
        if file_hash.lower() in known_hashes:
            continue

        task_row_index, task = tasks[task_id]
        result_text = read_document(path)
        result_id = f"RES-{task_id}-{version:02d}"

        criteria_rows = [
            (row_index, row)
            for row_index, row in iter_rows(criteria_ws)
            if str(row.get("task_id") or "") == task_id
        ]
        criteria_payload = [
            {
                "criterion_id": str(row.get("criterion_id") or ""),
                "criterion_text": str(
                    row.get("criterion_text") or ""
                ),
                "mandatory": (
                    str(row.get("mandatory")).lower()
                    not in {"false", "0", "no"}
                ),
            }
            for _, row in criteria_rows
        ]

        prompt = f"""
RESULT_EVALUATION_V4

Ты — AI-проверяющий проектного офиса. Оцени результат поручения по смыслу.
Верни ТОЛЬКО один JSON-объект без Markdown и комментариев.

Правила:
1. Не считать критерием простое совпадение слов.
2. Для каждого критерия указать статус Passed, Failed или Unclear и краткое
   доказательство из результата.
3. Не придумывать доказательства.
4. Если данных недостаточно для уверенной проверки, использовать Unclear.
5. В summary кратко объяснить управленческий вывод.
6. В rework_actions перечислить конкретные исправления.

Поручение:
{json.dumps({
    "task_id": task_id,
    "task_title": task.get("task_title"),
    "task_description": task.get("task_description"),
    "expected_result": task.get("expected_result"),
    "due_date": str(task.get("due_date") or ""),
}, ensure_ascii=False)}

Критерии:
{json.dumps(criteria_payload, ensure_ascii=False)}

JSON-схема:
{{
  "summary": "краткий вывод",
  "confidence": 0.0,
  "criteria": [
    {{
      "criterion_id": "точный ID из входа",
      "status": "Passed или Failed или Unclear",
      "evidence": "краткое доказательство"
    }}
  ],
  "rework_actions": ["конкретное действие"]
}}

РЕЗУЛЬТАТ:
---BEGIN---
{_bounded_text(result_text)}
---END---
""".strip()

        assessment = llm_json(
            purpose="result",
            prompt=prompt,
        )
        assessed_items = assessment.get("criteria")
        if not isinstance(assessed_items, list):
            raise LLMServiceError(
                f"LLM не вернула массив criteria для {path.name}."
            )
        assessed_by_id = {
            str(item.get("criterion_id") or ""): item
            for item in assessed_items
            if isinstance(item, dict)
            and item.get("criterion_id")
        }

        mandatory_statuses: list[str] = []
        passed_count = 0
        mandatory_count = 0

        for criterion_row_index, criterion in criteria_rows:
            criterion_id = str(
                criterion.get("criterion_id") or ""
            )
            assessed = assessed_by_id.get(criterion_id, {})
            status = str(
                assessed.get("status") or "Unclear"
            ).strip()
            if status not in {"Passed", "Failed", "Unclear"}:
                status = "Unclear"
            evidence = str(
                assessed.get("evidence")
                or "Модель не предоставила доказательство."
            ).strip()

            mandatory = (
                str(criterion.get("mandatory")).lower()
                not in {"false", "0", "no"}
            )
            if mandatory:
                mandatory_count += 1
                mandatory_statuses.append(status)
                if status == "Passed":
                    passed_count += 1

            update_mapping(
                criteria_ws,
                criterion_row_index,
                {
                    "status": status,
                    "evidence": evidence,
                    "result_id": result_id,
                    "checked_at": iso(now_moscow()),
                },
            )
        store.mark_dirty("03_task_register.xlsx")

        denominator = mandatory_count or len(criteria_rows) or 1
        score = round(passed_count / denominator * 100, 1)
        if "Failed" in mandatory_statuses:
            verdict = "Rework_Required"
        elif "Unclear" in mandatory_statuses:
            verdict = "Human_Review_Required"
        else:
            verdict = "Accepted"

        summary = str(
            assessment.get("summary")
            or (
                f"Выполнено обязательных критериев: "
                f"{passed_count} из {denominator}."
            )
        ).strip()
        rework_actions = [
            str(item).strip()
            for item in (
                assessment.get("rework_actions")
                if isinstance(assessment.get("rework_actions"), list)
                else []
            )
            if str(item).strip()
        ]
        if rework_actions:
            summary = (
                summary
                + " Действия: "
                + "; ".join(rework_actions)
            )[:1500]

        append_mapping(
            results_ws,
            {
                "result_id": result_id,
                "task_id": task_id,
                "version": version,
                "file_name": path.name,
                "file_path": str(path.resolve()),
                "file_hash": file_hash,
                "received_at": iso(now_moscow()),
                "status": "Checked",
                "acceptance_score": score,
                "verdict": verdict,
                "summary": summary,
                "rework_due_date": (
                    now_moscow().date() + timedelta(days=2)
                    if verdict == "Rework_Required"
                    else None
                ),
                "checked_at": iso(now_moscow()),
                "idempotency_key": file_hash,
            },
        )
        store.mark_dirty("03_task_register.xlsx")

        append_mapping(
            inbox_ws,
            {
                "inbox_id": f"INBOX-{uuid.uuid4().hex[:10].upper()}",
                "task_id": task_id,
                "version": version,
                "file_name": path.name,
                "file_path": str(path.resolve()),
                "file_hash": file_hash,
                "available_from": iso(now_moscow()),
                "received_at": iso(now_moscow()),
                "status": "Processed",
                "run_id": run_id,
                "result_id": result_id,
                "idempotency_key": file_hash,
                "error": "",
            },
        )
        store.mark_dirty("09_source_register.xlsx")
        known_hashes.add(file_hash.lower())

        old_status = str(task.get("status") or "")
        new_status = {
            "Accepted": "Closed",
            "Rework_Required": "Rework_Required",
            "Human_Review_Required": "Human_Review_Required",
        }[verdict]

        update_mapping(
            tasks_ws,
            task_row_index,
            {
                "status": new_status,
                "timing_state": (
                    "Closed"
                    if verdict == "Accepted"
                    else timing_state(
                        parse_date(task.get("due_date")),
                        new_status,
                        now_moscow().date(),
                    )
                ),
                "acceptance_status": verdict,
                "current_iteration": version,
                "rework_due_date": (
                    now_moscow().date() + timedelta(days=2)
                    if verdict == "Rework_Required"
                    else None
                ),
                "latest_result_id": result_id,
                "last_status_at": iso(now_moscow()),
                "closed_at": (
                    iso(now_moscow())
                    if verdict == "Accepted"
                    else None
                ),
                "updated_at": iso(now_moscow()),
            },
        )
        store.mark_dirty("03_task_register.xlsx")

        append_history(
            store,
            task_id=task_id,
            run_id=run_id,
            from_status=old_status,
            to_status=new_status,
            reason="Результат проверен языковой моделью",
            skill_id="SK-13",
            details=summary,
        )
        append_action_log(
            store,
            run_id=run_id,
            skill_id="SK-13",
            action_type="llm_check_result",
            object_type="result",
            object_id=result_id,
            project_id=str(task.get("project_id") or ""),
            source_file=path.name,
            target_file="03_task_register.xlsx",
            target_sheet="Task_Results",
            details=f"{verdict}; score={score}; {summary[:500]}",
        )

        metrics["results_checked"] += 1
        if verdict == "Accepted":
            metrics["accepted"] += 1
            metrics["tasks_closed"] += 1
        elif verdict == "Rework_Required":
            metrics["rework_required"] += 1
        else:
            metrics["human_review_required"] += 1

    return metrics


def create_notification(
    store: WorkbookStore,
    task: dict[str, Any],
    notification_type: str,
    requested_action: str,
    due_date: date | None,
    escalation_level: int,
) -> bool:
    workbook = store.open("05_message_outbox.xlsx")
    ws = workbook["Outbox"]
    task_id = str(task.get("task_id") or "")
    key = hashlib.sha256(
        f"{task_id}|{notification_type}|{due_date}".encode("utf-8")
    ).hexdigest()
    for _, row in iter_rows(ws):
        if str(row.get("idempotency_key") or "") == key:
            return False

    timestamp = now_moscow()
    append_mapping(
        ws,
        {
            "message_id": f"MSG-{timestamp:%Y%m%d%H%M%S}-{uuid.uuid4().hex[:5]}",
            "created_at": iso(timestamp),
            "notification_type": notification_type,
            "task_id": task_id,
            "project_id": task.get("project_id"),
            "recipient_id": task.get("assignee_id"),
            "recipient_name": task.get("assignee_name"),
            "recipient_email": task.get("assignee_email"),
            "subject": f"{notification_type}: {task_id}",
            "facts": (
                f"Поручение: {task.get('task_title')}. "
                f"Срок: {task.get('due_date')}."
            ),
            "requested_action": requested_action,
            "action_due_date": due_date,
            "channel": "Local_Draft",
            "status": "Draft",
            "delivery_status": "Not_Sent",
            "escalation_level": escalation_level,
            "iteration": task.get("current_iteration") or 0,
            "idempotency_key": key,
            "approved_by": "",
            "sent_at": None,
        },
    )
    store.mark_dirty("05_message_outbox.xlsx")
    return True


def check_deadlines(
    store: WorkbookStore,
    run_id: str,
) -> dict[str, int]:
    metrics = {
        "due_soon": 0,
        "overdue": 0,
        "messages_created": 0,
    }
    workbook = store.open("03_task_register.xlsx")
    ws = workbook["Tasks"]
    today = now_moscow().date()

    for row_index, task in iter_rows(ws):
        status = str(task.get("status") or "")
        due = parse_date(task.get("due_date"))
        new_timing = timing_state(due, status, today)
        old_timing = str(task.get("timing_state") or "")
        escalation = int(task.get("escalation_level") or 0)

        if new_timing in {"Due_Soon", "Due_Today"}:
            metrics["due_soon"] += 1
            if create_notification(
                store,
                task,
                "Due_Soon",
                "Подтвердить план выполнения поручения.",
                due,
                escalation,
            ):
                metrics["messages_created"] += 1
        elif new_timing == "Overdue":
            metrics["overdue"] += 1
            escalation = max(escalation, 1)
            if create_notification(
                store,
                task,
                "Overdue",
                "Предоставить статус и план восстановления срока.",
                today,
                escalation,
            ):
                metrics["messages_created"] += 1

        if new_timing != old_timing or escalation != int(
            task.get("escalation_level") or 0
        ):
            update_mapping(
                ws,
                row_index,
                {
                    "timing_state": new_timing,
                    "escalation_level": escalation,
                    "last_reminder_at": (
                        iso(now_moscow())
                        if new_timing in {"Due_Soon", "Due_Today", "Overdue"}
                        else task.get("last_reminder_at")
                    ),
                    "updated_at": iso(now_moscow()),
                },
            )
            store.mark_dirty("03_task_register.xlsx")

        append_action_log(
            store,
            run_id=run_id,
            skill_id="SK-14",
            action_type="check_deadline",
            object_type="task",
            object_id=str(task.get("task_id") or ""),
            project_id=str(task.get("project_id") or ""),
            target_file="03_task_register.xlsx",
            target_sheet="Tasks",
            details=f"{old_timing or 'None'} → {new_timing}",
        )

    return metrics


def update_skill_state(
    store: WorkbookStore,
    *,
    skill_id: str,
    status: str,
    processed: int,
    blocked: int = 0,
    error: str = "",
) -> None:
    workbook = store.open("07_task_control_state.xlsx")
    ws = workbook["Skill_State"]
    found = find_row(ws, "skill_id", skill_id)
    if not found:
        return
    row_index, _ = found
    update_mapping(
        ws,
        row_index,
        {
            "last_started_at": iso(now_moscow()),
            "last_completed_at": iso(now_moscow()),
            "last_status": status,
            "processed_objects": processed,
            "blocked_objects": blocked,
            "action_log_rows": processed,
            "last_error": error,
        },
    )
    store.mark_dirty("07_task_control_state.xlsx")


def update_schedule_state(
    store: WorkbookStore,
    *,
    run_id: str,
    trigger_type: str,
    error: str = "",
) -> None:
    workbook = store.open("10_loop_trigger_settings.xlsx")
    ws = workbook["Schedule_State"]
    next_run = next_three_hour_boundary(now_moscow())
    values = {
        "last_dispatch_at": iso(now_moscow()),
        "last_trigger_type": trigger_type,
        "last_trigger_id": run_id,
        "last_run_id": run_id,
        "last_error": error,
    }
    if trigger_type in {"periodic", "widget_full", "watcher_full"}:
        values["last_full_sweep_at"] = iso(now_moscow())
    values["next_full_sweep_at"] = iso(next_run)

    for row_index, row in iter_rows(ws):
        name = str(row.get("state_name") or "")
        if name in values:
            update_mapping(
                ws,
                row_index,
                {
                    "state_value": values[name],
                    "updated_at": iso(now_moscow()),
                },
            )
    store.mark_dirty("10_loop_trigger_settings.xlsx")


def next_three_hour_boundary(now: datetime) -> datetime:
    local = now.astimezone(MOSCOW).replace(second=0, microsecond=0)
    next_hour = ((local.hour // 3) + 1) * 3
    if next_hour >= 24:
        return local.replace(hour=0, minute=0) + timedelta(days=1)
    return local.replace(hour=next_hour, minute=0)


def update_dashboard(root: Path, store: WorkbookStore) -> None:
    path = root / "output" / "dashboards" / "10_task_dashboard.xlsx"
    if not path.exists():
        return

    task_book = store.open("03_task_register.xlsx")
    tasks = [row for _, row in iter_rows(task_book["Tasks"])]
    source_book = store.open("09_source_register.xlsx")
    sources = [row for _, row in iter_rows(source_book["Source_Documents"])]
    results = [row for _, row in iter_rows(task_book["Task_Results"])]
    action_book = store.open("08_skill_action_log.xlsx")
    actions = [row for _, row in iter_rows(action_book["Action_Log"])]

    workbook = load_workbook(path)
    try:
        if "Executive_Dashboard" in workbook.sheetnames:
            ws = workbook["Executive_Dashboard"]
            active = sum(
                1 for task in tasks
                if str(task.get("status") or "") not in {"Closed", "Cancelled"}
            )
            due_soon = sum(
                1 for task in tasks
                if str(task.get("timing_state") or "")
                in {"Due_Soon", "Due_Today"}
            )
            overdue = sum(
                1 for task in tasks
                if str(task.get("timing_state") or "") == "Overdue"
            )
            validating = sum(
                1 for task in tasks
                if str(task.get("status") or "")
                in {"Result_Submitted", "Validating"}
            )
            rework = sum(
                1 for task in tasks
                if str(task.get("status") or "") == "Rework_Required"
            )
            closed = sum(
                1 for task in tasks
                if str(task.get("status") or "") == "Closed"
            )
            ws["A5"] = len(tasks)
            ws["C5"] = active
            ws["E5"] = due_soon
            ws["G5"] = overdue
            ws["I5"] = validating
            ws["K5"] = rework
            ws["M5"] = closed
            ws["A10"] = len(sources)
            ws["C10"] = len(tasks)
            ws["E10"] = len(tasks)
            ws["G10"] = len(results)
            ws["I10"] = sum(
                1 for row in results if str(row.get("verdict")) == "Accepted"
            )
            ws["K10"] = sum(
                1 for row in results
                if str(row.get("verdict")) == "Rework_Required"
            )
            ws["M10"] = closed

            critical = sorted(
                [
                    task for task in tasks
                    if str(task.get("timing_state") or "")
                    in {"Due_Soon", "Due_Today", "Overdue"}
                ],
                key=lambda row: (
                    0 if str(row.get("timing_state")) == "Overdue" else 1,
                    str(row.get("due_date") or ""),
                ),
            )[:8]
            for row in range(15, 23):
                for col in range(1, 15):
                    ws.cell(row, col).value = None
            for offset, task in enumerate(critical, start=15):
                ws.cell(offset, 1).value = task.get("task_id")
                ws.cell(offset, 2).value = task.get("task_title")
                ws.cell(offset, 3).value = task.get("assignee_name")
                ws.cell(offset, 4).value = task.get("due_date")
                ws.cell(offset, 5).value = task.get("timing_state")
                ws.cell(offset, 6).value = task.get("escalation_level")
                ws.cell(offset, 7).value = (
                    "Предоставить статус"
                    if str(task.get("timing_state")) == "Overdue"
                    else "Подтвердить план"
                )
            for offset, action in enumerate(
                sorted(
                    actions,
                    key=lambda row: str(row.get("logged_at") or ""),
                    reverse=True,
                )[:8],
                start=15,
            ):
                ws.cell(offset, 8).value = action.get("logged_at")
                ws.cell(offset, 9).value = action.get("run_id")
                ws.cell(offset, 10).value = action.get("skill_id")
                ws.cell(offset, 11).value = action.get("action_type")
                ws.cell(offset, 12).value = action.get("object_id")
                ws.cell(offset, 13).value = action.get("action_status")
                ws.cell(offset, 14).value = action.get("details")

        if "Task_Dashboard_Data" in workbook.sheetnames:
            ws = workbook["Task_Dashboard_Data"]
            project_book = store.open("01_projects.xlsx")
            projects = [row for _, row in iter_rows(project_book["Projects"])]
            candidates = [
                row for _, row in iter_rows(source_book["Extracted_Candidates"])
            ]
            for row_index in range(4, ws.max_row + 1):
                project_id = str(ws.cell(row_index, 1).value or "")
                if not project_id:
                    continue
                project_tasks = [
                    task for task in tasks
                    if str(task.get("project_id") or "") == project_id
                ]
                project_candidates = [
                    item for item in candidates
                    if str(item.get("project_id") or "") == project_id
                ]
                ws.cell(row_index, 3).value = len(project_candidates)
                ws.cell(row_index, 4).value = len(project_tasks)
                ws.cell(row_index, 5).value = sum(
                    1 for task in project_tasks
                    if str(task.get("status") or "") not in {"Closed", "Cancelled"}
                )
                ws.cell(row_index, 6).value = sum(
                    1 for task in project_tasks
                    if str(task.get("timing_state") or "")
                    in {"Due_Soon", "Due_Today"}
                )
                ws.cell(row_index, 7).value = sum(
                    1 for task in project_tasks
                    if str(task.get("timing_state") or "") == "Overdue"
                )
                ws.cell(row_index, 8).value = sum(
                    1 for task in project_tasks
                    if str(task.get("status") or "") in {"Validating", "Result_Submitted"}
                )
                ws.cell(row_index, 9).value = sum(
                    1 for task in project_tasks
                    if str(task.get("status") or "") == "Rework_Required"
                )
                ws.cell(row_index, 10).value = sum(
                    1 for task in project_tasks
                    if str(task.get("status") or "") == "Closed"
                )
                ws.cell(row_index, 11).value = iso(now_moscow())
                ws.cell(row_index, 12).value = iso(now_moscow())
                ws.cell(row_index, 13).value = "Ready"

        temp = path.with_name(
            f".{path.stem}.{uuid.uuid4().hex[:8]}.tmp.xlsx"
        )
        workbook.save(temp)
        os.replace(temp, path)
    finally:
        workbook.close()


def update_control_state(
    store: WorkbookStore,
    *,
    run_id: str,
    trigger_type: str,
    started_at: datetime,
    completed_at: datetime,
    status: str,
    metrics: dict[str, int],
    error: str,
) -> None:
    workbook = store.open("07_task_control_state.xlsx")
    control_ws = workbook["Control_State"]
    found = find_row(control_ws, "control_id", "task_lifecycle_pmo")
    if found:
        row_index, row = found
        update_mapping(
            control_ws,
            row_index,
            {
                "last_started_at": iso(started_at),
                "last_completed_at": iso(completed_at),
                "last_status": status,
                "last_success_at": (
                    iso(completed_at)
                    if status in {"Success", "Skipped"}
                    else row.get("last_success_at")
                ),
                "sources_processed": metrics.get("sources_processed", 0),
                "candidates_ready": metrics.get("assignments_found", 0),
                "tasks_registered": metrics.get("tasks_registered", 0),
                "tasks_due_soon": metrics.get("due_soon", 0),
                "tasks_overdue": metrics.get("overdue", 0),
                "results_checked": metrics.get("results_checked", 0),
                "tasks_closed": metrics.get("tasks_closed", 0),
                "action_log_rows": metrics.get("action_log_rows", 0),
                "action_log_status": "Ready",
                "active_run_id": "",
                "last_error": error,
            },
        )

    run_ws = workbook["Run_Log"]
    append_mapping(
        run_ws,
        {
            "run_id": run_id,
            "control_id": "task_lifecycle_pmo",
            "trigger_type": trigger_type,
            "trigger_id": run_id,
            "started_at": iso(started_at),
            "completed_at": iso(completed_at),
            "status": status,
            "virtual_now": None,
            "sources_processed": metrics.get("sources_processed", 0),
            "assignments_found": metrics.get("assignments_found", 0),
            "tasks_registered": metrics.get("tasks_registered", 0),
            "due_soon": metrics.get("due_soon", 0),
            "overdue": metrics.get("overdue", 0),
            "results_checked": metrics.get("results_checked", 0),
            "accepted": metrics.get("accepted", 0),
            "rework_required": metrics.get("rework_required", 0),
            "tasks_closed": metrics.get("tasks_closed", 0),
            "messages_created": metrics.get("messages_created", 0),
            "dashboard_updated": True,
            "action_log_rows": metrics.get("action_log_rows", 0),
            "error_summary": error,
        },
    )

    observability_ws = workbook["Loop_Observability"]
    stage_metrics = (
        ("SK-18", metrics.get("sources_processed", 0), metrics.get("assignments_found", 0)),
        ("SK-12", metrics.get("assignments_found", 0), metrics.get("tasks_registered", 0)),
        ("SK-13", metrics.get("results_checked", 0), metrics.get("tasks_closed", 0)),
        ("SK-14", metrics.get("tasks_registered", 0), metrics.get("due_soon", 0) + metrics.get("overdue", 0)),
    )
    for order, (component, input_count, output_count) in enumerate(stage_metrics, start=1):
        append_mapping(
            observability_ws,
            {
                "run_id": run_id,
                "step_order": order,
                "component_id": component,
                "started_at": iso(started_at),
                "completed_at": iso(completed_at),
                "duration_ms": int((completed_at - started_at).total_seconds() * 1000),
                "input_count": input_count,
                "output_count": output_count,
                "status": status,
                "details": "Hybrid LLM runtime v4: semantic AI + local Excel orchestrator",
            },
        )

    store.mark_dirty("07_task_control_state.xlsx")


def move_command(
    root: Path,
    command_path: Path | None,
    success: bool,
) -> None:
    if command_path is None or not command_path.exists():
        return
    target_dir = (
        root / "archive" / "commands" / ("processed" if success else "failed")
    )
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / command_path.name
    if target.exists():
        target = target_dir / f"{command_path.stem}-{uuid.uuid4().hex[:5]}.json"
    command_path.replace(target)


def count_action_rows(store: WorkbookStore) -> int:
    workbook = store.open("08_skill_action_log.xlsx")
    return sum(1 for _ in iter_rows(workbook["Action_Log"]))


def run(
    action: str,
    *,
    root: Path | None = None,
    trigger_type: str = "widget",
    trigger_id: str = "",
    command_path: Path | None = None,
) -> dict[str, Any]:
    root = (root or find_project_root()).resolve()
    ensure_structure(root)

    allowed_actions = {
        "run_full_control",
        "process_meetings",
        "check_results",
        "check_deadlines",
        "refresh_dashboard",
    }
    if action not in allowed_actions:
        return {
            "ok": False,
            "message": f"Неизвестное действие: {action}",
        }

    started_at = now_moscow()
    run_id = (
        f"RUN-{started_at:%Y%m%d-%H%M%S}-{uuid.uuid4().hex[:5].upper()}"
    )
    metrics = {
        "sources_processed": 0,
        "assignments_found": 0,
        "tasks_registered": 0,
        "results_checked": 0,
        "accepted": 0,
        "rework_required": 0,
        "tasks_closed": 0,
        "due_soon": 0,
        "overdue": 0,
        "messages_created": 0,
        "action_log_rows": 0,
    }
    store = WorkbookStore(root)
    success = False
    error = ""

    try:
        with processor_lock(root):
            settings = load_settings(store)
            limit = max(1, int(settings.get("max_items_per_run") or 100))

            if action in {"run_full_control", "process_meetings"}:
                result = process_meetings(store, root, run_id, limit)
                metrics.update({
                    key: metrics.get(key, 0) + value
                    for key, value in result.items()
                })

            if action in {"run_full_control", "check_results"}:
                result = process_results(store, root, run_id, limit)
                metrics.update({
                    key: metrics.get(key, 0) + value
                    for key, value in result.items()
                })

            if action in {"run_full_control", "check_deadlines"}:
                result = check_deadlines(store, run_id)
                metrics.update({
                    key: metrics.get(key, 0) + value
                    for key, value in result.items()
                })

            update_skill_state(
                store,
                skill_id="SK-18",
                status="Success",
                processed=metrics["sources_processed"],
            )
            update_skill_state(
                store,
                skill_id="SK-12",
                status="Success",
                processed=metrics["tasks_registered"],
            )
            update_skill_state(
                store,
                skill_id="SK-13",
                status="Success",
                processed=metrics["results_checked"],
            )
            update_skill_state(
                store,
                skill_id="SK-14",
                status="Success",
                processed=metrics["due_soon"] + metrics["overdue"],
            )
            update_skill_state(
                store,
                skill_id="L-09",
                status="Success",
                processed=sum(metrics.values()),
            )

            metrics["action_log_rows"] = count_action_rows(store)
            completed_at = now_moscow()
            run_status = (
                "Skipped"
                if not any(
                    metrics[key]
                    for key in (
                        "sources_processed",
                        "tasks_registered",
                        "results_checked",
                        "due_soon",
                        "overdue",
                        "messages_created",
                    )
                )
                else "Success"
            )
            update_control_state(
                store,
                run_id=run_id,
                trigger_type=trigger_type,
                started_at=started_at,
                completed_at=completed_at,
                status=run_status,
                metrics=metrics,
                error="",
            )
            update_schedule_state(
                store,
                run_id=run_id,
                trigger_type=trigger_type,
                error="",
            )
            store.save_all()
            update_dashboard(root, store)
            success = True
            move_command(root, command_path, True)

            return {
                "ok": True,
                "run_id": run_id,
                "status": run_status,
                "message": (
                    "Контроль выполнен с языковой моделью. "
                    f"Новых поручений: {metrics['tasks_registered']}; "
                    f"результатов проверено: {metrics['results_checked']}; "
                    f"закрыто: {metrics['tasks_closed']}."
                ),
                "metrics": metrics,
                "project_root": str(root),
            }
    except Exception as exc:
        error = str(exc)
        completed_at = now_moscow()
        try:
            append_action_log(
                store,
                run_id=run_id,
                skill_id="L-09",
                action_type="run_failed",
                object_type="run",
                object_id=run_id,
                status="Failed",
                details=error,
                error_code=type(exc).__name__,
            )
            update_skill_state(
                store,
                skill_id="L-09",
                status="Blocked",
                processed=0,
                blocked=1,
                error=error,
            )
            update_control_state(
                store,
                run_id=run_id,
                trigger_type=trigger_type,
                started_at=started_at,
                completed_at=completed_at,
                status="Blocked",
                metrics=metrics,
                error=error,
            )
            update_schedule_state(
                store,
                run_id=run_id,
                trigger_type=trigger_type,
                error=error,
            )
            store.save_all()
        except Exception:
            pass
        move_command(root, command_path, False)
        return {
            "ok": False,
            "run_id": run_id,
            "status": "Blocked",
            "message": error,
            "metrics": metrics,
            "project_root": str(root),
        }
    finally:
        store.close()


def read_runtime_state(root: Path | None = None) -> dict[str, Any]:
    root = (root or find_project_root()).resolve()
    path = root / "archive" / ".local_runtime_state.json"
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {
        "status": "not_started",
        "heartbeat_at": None,
        "last_run_at": None,
        "last_error": None,
        "project_root": str(root),
    }


def write_runtime_state(root: Path, state: dict[str, Any]) -> None:
    path = root / "archive" / ".local_runtime_state.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(".tmp")
    temp.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temp.replace(path)
