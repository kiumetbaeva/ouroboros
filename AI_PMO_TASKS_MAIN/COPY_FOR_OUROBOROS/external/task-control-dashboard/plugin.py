from __future__ import annotations

import importlib.util
import json
import os
import sys
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Ouroboros loads extension modules by file path and does not guarantee that
# the skill directory is present in sys.path. Load the sibling runtime.py
# explicitly from the reviewed skill payload.
_RUNTIME_PATH = Path(__file__).resolve().with_name("runtime.py")
_RUNTIME_SPEC = importlib.util.spec_from_file_location(
    "ai_pmo_task_control_runtime_v410",
    _RUNTIME_PATH,
)
if _RUNTIME_SPEC is None or _RUNTIME_SPEC.loader is None:
    raise ModuleNotFoundError(
        f"Не удалось подготовить загрузку локального runtime: {_RUNTIME_PATH}"
    )
_RUNTIME = importlib.util.module_from_spec(_RUNTIME_SPEC)
sys.modules[_RUNTIME_SPEC.name] = _RUNTIME
_RUNTIME_SPEC.loader.exec_module(_RUNTIME)

PROJECT_FOLDER = _RUNTIME.PROJECT_FOLDER
find_project_root = _RUNTIME.find_project_root
iso = _RUNTIME.iso
next_three_hour_boundary = _RUNTIME.next_three_hour_boundary
now_moscow = _RUNTIME.now_moscow
parse_dt = _RUNTIME.parse_dt
read_runtime_state = _RUNTIME.read_runtime_state
run = _RUNTIME.run
sha256_file = _RUNTIME.sha256_file


SKILL_NAME = "task-control-dashboard"
MOSCOW = timezone(timedelta(hours=3))

STATUS_RU = {
    "Success": "Выполнено успешно",
    "Partial": "Выполнено частично",
    "Blocked": "Выполнение заблокировано",
    "Skipped": "Новых данных нет",
    "Never": "Ещё не запускался",
    "Registered": "Зарегистрировано",
    "In_Progress": "В работе",
    "In Progress": "В работе",
    "Result_Submitted": "Результат предоставлен",
    "Validating": "Результат проверяется",
    "Rework_Required": "Требуется доработка",
    "Rework Required": "Требуется доработка",
    "Human_Review_Required": "Требуется ручная проверка",
    "Human Review Required": "Требуется ручная проверка",
    "Accepted": "Результат принят",
    "Closed": "Закрыто",
    "Cancelled": "Отменено",
    "On_Track": "Срок в норме",
    "On Track": "Срок в норме",
    "Due_Soon": "Срок приближается",
    "Due Soon": "Срок приближается",
    "Due_Today": "Срок сегодня",
    "Due Today": "Срок сегодня",
    "Overdue": "Просрочено",
    "Draft": "Подготовлено к отправке",
    "New": "Новое",
    "Processed": "Обработано",
    "Failed": "Ошибка",
    "Queued": "В очереди",
    None: "Нет данных",
    "": "Нет данных",
}

ACTION_NAMES = {
    "run_full_control": "Полный контроль поручений",
    "process_meetings": "Обработка новых протоколов",
    "check_deadlines": "Проверка сроков и просрочек",
    "check_results": "Проверка результатов",
}

PROFILE_LABELS = {
    "HYBRID": "AI-контур: кнопки + новые файлы + каждые 3 часа",
    "WIDGET_ONLY": "Только ручной запуск из виджета",
    "FILES_IMMEDIATELY": "Кнопки + новые файлы",
    "EVERY_3_HOURS": "Кнопки + полный контроль каждые 3 часа",
}


def ru(value: Any) -> str:
    raw = None if value is None else str(value)
    return STATUS_RU.get(raw, raw or "Нет данных")


def sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    last_error = None
    for attempt in range(4):
        workbook = None
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                return []
            ws = workbook[sheet_name]
            names = [str(cell.value or "").strip() for cell in ws[3]]
            result = []
            for values in ws.iter_rows(
                min_row=4,
                max_col=max(1, len(names)),
                values_only=True,
            ):
                values = list(values)
                if not any(value not in (None, "") for value in values):
                    continue
                result.append(
                    {
                        name: values[index] if index < len(values) else None
                        for index, name in enumerate(names)
                        if name
                    }
                )
            return result
        except Exception as exc:
            last_error = exc
            if attempt < 3:
                import time
                time.sleep(0.15 * (attempt + 1))
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass
    raise RuntimeError(f"Не удалось прочитать {path.name}: {last_error}")


def registry_mtime(root: Path) -> float:
    newest = 0.0
    for name in (
        "03_task_register.xlsx",
        "05_message_outbox.xlsx",
        "07_task_control_state.xlsx",
        "08_skill_action_log.xlsx",
        "09_source_register.xlsx",
        "10_loop_trigger_settings.xlsx",
    ):
        try:
            newest = max(
                newest,
                (root / "registry" / name).stat().st_mtime,
            )
        except OSError:
            pass
    return newest


def data_revision(root: Path) -> str:
    latest = registry_mtime(root)
    return str(int(latest * 1000)) if latest else "0"


def registry_updated_at(root: Path) -> str | None:
    latest = registry_mtime(root)
    if not latest:
        return None
    return datetime.fromtimestamp(latest, MOSCOW).isoformat()


def load_config(root: Path) -> dict[str, Any]:
    config = {
        "active_profile": "HYBRID",
        "widget_commands_enabled": True,
        "file_events_enabled": True,
        "periodic_full_run_enabled": True,
        "watcher_poll_seconds": 2,
        "periodic_interval_minutes": 180,
        "debounce_seconds": 15,
    }
    for row in sheet_rows(
        root / "registry" / "10_loop_trigger_settings.xlsx",
        "Control_Panel",
    ):
        key = str(row.get("setting_name") or "").strip()
        if not key:
            continue
        value = row.get("setting_value")
        value_type = str(row.get("value_type") or "Text")
        if value_type == "Boolean":
            config[key] = str(value).strip().lower() in {
                "true", "1", "yes", "да"
            }
        elif value_type == "Integer":
            try:
                config[key] = int(float(value))
            except Exception:
                pass
        else:
            config[key] = str(value or "").strip()
    config["active_profile"] = str(
        config.get("active_profile") or "HYBRID"
    ).upper()
    return config


def is_ignored_input_file(path: Path) -> bool:
    name = path.name.strip().lower()
    return (
        not path.is_file()
        or name.startswith("readme")
        or name.startswith(".")
        or name.startswith("00_")
        or name.startswith("класть_")
    )


def count_files(path: Path, suffixes: set[str] | None = None) -> int:
    if not path.exists():
        return 0
    count = 0
    for item in path.iterdir():
        if is_ignored_input_file(item):
            continue
        if suffixes and item.suffix.lower() not in suffixes:
            continue
        count += 1
    return count


def pending_input_count(
    root: Path,
    folder: str,
    registry_sheet: str,
) -> int:
    source_path = root / "registry" / "09_source_register.xlsx"
    known = {
        str(row.get("file_hash") or "").lower()
        for row in sheet_rows(source_path, registry_sheet)
        if row.get("file_hash")
    }
    count = 0
    for path in (root / "input" / folder).iterdir():
        if (
            is_ignored_input_file(path)
            or path.suffix.lower() not in {".txt", ".md", ".docx", ".pdf"}
        ):
            continue
        try:
            if sha256_file(path).lower() not in known:
                count += 1
        except OSError:
            pass
    return count


def format_activity(row: dict[str, Any]) -> dict[str, Any]:
    skill_id = str(row.get("skill_id") or "")
    action = str(row.get("action_type") or "Действие")
    object_id = str(row.get("object_id") or "")
    details = str(row.get("details") or "")
    title = {
        "SK-18": "Обработан протокол встречи",
        "SK-12": "Зарегистрировано поручение",
        "SK-14": "Проверены сроки",
        "SK-13": "Проверен результат",
        "L-09": "Выполнен AI-контур",
    }.get(skill_id, action.replace("_", " ").capitalize())
    if object_id:
        details = f"{object_id}. {details}".strip()
    return {
        "time": iso(parse_dt(row.get("logged_at"))),
        "title": title,
        "details": details[:240],
        "status": ru(row.get("action_status")),
        "tone": (
            "danger"
            if row.get("error_code")
            else "warning"
            if "Overdue" in str(row) or "Rework" in str(row)
            else "success"
        ),
    }


def schedule_label(config: dict[str, Any]) -> str:
    profile = str(config.get("active_profile") or "HYBRID")
    poll = int(config.get("watcher_poll_seconds") or 2)
    interval = int(config.get("periodic_interval_minutes") or 180)
    hours = interval / 60
    interval_text = (
        f"{int(hours)} ч." if hours.is_integer() else f"{interval} мин."
    )
    if profile == "WIDGET_ONLY":
        return "обработка запускается только кнопками"
    if profile == "FILES_IMMEDIATELY":
        return f"кнопки и новые файлы; проверка файлов каждые {poll} сек."
    if profile == "EVERY_3_HOURS":
        return f"кнопки и полный контроль каждые {interval_text}"
    return (
        f"кнопки и новые файлы; локальная проверка каждые {poll} сек.; "
        f"полный контроль каждые {interval_text}"
    )


def read_watcher_heartbeat(root: Path) -> dict[str, Any]:
    path = root / "archive" / ".local_runtime_heartbeat.json"
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}


def get_dashboard_data() -> dict[str, Any]:
    errors: list[str] = []
    try:
        root = find_project_root()
    except Exception as exc:
        return {
            "ok": False,
            "project_root": "",
            "updated_at": iso(now_moscow()),
            "header": {
                "health": "Рабочая папка не установлена",
                "next_run": iso(next_three_hour_boundary(now_moscow())),
                "seconds_to_next": None,
                "full_sweep_due": False,
                "last_status": "Ошибка",
                "last_run_id": "",
            },
            "kpi": {},
            "stages": [],
            "tasks": [],
            "activity": [],
            "errors": [str(exc)],
            "system": {
                "storage_ready": False,
                "data_revision": "0",
                "registry_updated_at": None,
                "workspace_candidates": 0,
                "pending_commands": 0,
                "meeting_files": 0,
                "result_files": 0,
                "active_profile": "HYBRID",
                "profile_label": PROFILE_LABELS["HYBRID"],
                "schedule_label": "Рабочая папка не установлена",
                "watcher_status": "not_started",
            },
            "paths": {},
        }

    registry = root / "registry"
    required = [
        "03_task_register.xlsx",
        "05_message_outbox.xlsx",
        "07_task_control_state.xlsx",
        "08_skill_action_log.xlsx",
        "09_source_register.xlsx",
        "10_loop_trigger_settings.xlsx",
    ]
    missing = [name for name in required if not (registry / name).is_file()]
    if missing:
        errors.append("Не найдены реестры: " + ", ".join(missing))

    try:
        config = load_config(root)
        tasks = sheet_rows(registry / "03_task_register.xlsx", "Tasks")
        results = sheet_rows(
            registry / "03_task_register.xlsx",
            "Task_Results",
        )
        outbox = sheet_rows(registry / "05_message_outbox.xlsx", "Outbox")
        action_rows = sheet_rows(
            registry / "08_skill_action_log.xlsx",
            "Action_Log",
        )
        control_rows = sheet_rows(
            registry / "07_task_control_state.xlsx",
            "Control_State",
        )
        run_rows = sheet_rows(
            registry / "07_task_control_state.xlsx",
            "Run_Log",
        )
        skill_rows = sheet_rows(
            registry / "07_task_control_state.xlsx",
            "Skill_State",
        )
    except Exception as exc:
        errors.append(str(exc))
        config = {"active_profile": "HYBRID"}
        tasks, results, outbox, action_rows = [], [], [], []
        control_rows, run_rows, skill_rows = [], [], []

    control = control_rows[0] if control_rows else {}
    last_run = run_rows[-1] if run_rows else {}
    last_status_raw = control.get("last_status") or last_run.get("status")
    if (
        control.get("last_error")
        and str(last_status_raw or "") not in {"Success", "Skipped"}
    ):
        errors.append(str(control.get("last_error")))

    runtime_state = read_runtime_state(root)
    heartbeat_state = read_watcher_heartbeat(root)
    heartbeat = parse_dt(
        heartbeat_state.get("heartbeat_at")
        or runtime_state.get("heartbeat_at")
    )
    heartbeat_age = None
    if heartbeat is not None:
        heartbeat_age = max(
            0,
            int((now_moscow() - heartbeat).total_seconds()),
        )
    watcher_status = str(
        heartbeat_state.get("status")
        or runtime_state.get("status")
        or "not_started"
    )
    watcher_running = (
        watcher_status == "running"
        and heartbeat_age is not None
        and heartbeat_age <= 45
    )
    profile = str(config.get("active_profile") or "HYBRID").upper()
    watcher_required = profile != "WIDGET_ONLY"
    if watcher_required and not watcher_running:
        if watcher_status in {"error", "stopped"} or (
            watcher_status == "running"
            and heartbeat_age is not None
            and heartbeat_age > 90
        ):
            errors.append(
                "Локальный watcher остановлен или heartbeat устарел. "
                "Кнопки работают, но автоматический запуск по файлам и "
                "расписанию недоступен."
            )

    def count_status(values: set[str]) -> int:
        return sum(
            1 for task in tasks
            if str(task.get("status") or "") in values
        )

    def count_timing(values: set[str]) -> int:
        return sum(
            1 for task in tasks
            if str(task.get("timing_state") or "") in values
        )

    latest_tasks = sorted(
        tasks,
        key=lambda row: (
            parse_dt(row.get("updated_at") or row.get("created_at"))
            or datetime.min.replace(tzinfo=MOSCOW)
        ),
        reverse=True,
    )[:20]
    task_items = []
    for task in latest_tasks:
        problem = ""
        if str(task.get("status")) == "Rework_Required":
            problem = "Часть критериев не выполнена"
        elif str(task.get("status")) == "Human_Review_Required":
            problem = "Нужна проверка человеком"
        elif str(task.get("timing_state")) == "Overdue":
            problem = "Срок поручения нарушен"
        task_items.append(
            {
                "id": task.get("task_id"),
                "title": (
                    task.get("task_title")
                    or task.get("task_description")
                    or "Без названия"
                ),
                "assignee": task.get("assignee_name") or "Не определён",
                "due_date": iso(parse_dt(task.get("due_date"))),
                "status": ru(task.get("status")),
                "timing": ru(task.get("timing_state")),
                "problem": problem,
                "project_id": task.get("project_id"),
            }
        )

    skill_map = {
        str(row.get("skill_id")): row
        for row in skill_rows
        if row.get("skill_id")
    }
    stage_items = []
    for skill_id, name in (
        ("SK-18", "Разбор протоколов"),
        ("SK-12", "Регистрация поручений"),
        ("SK-14", "Контроль сроков"),
        ("SK-13", "Проверка результатов"),
    ):
        row = skill_map.get(skill_id, {})
        processed = int(row.get("processed_objects") or 0)
        blocked = int(row.get("blocked_objects") or 0)
        raw_status = str(row.get("last_status") or "Never")
        tone = (
            "warning"
            if blocked or raw_status in {"Blocked", "Partial"}
            else "success"
            if raw_status in {"Success", "Skipped"}
            else "neutral"
        )
        stage_items.append(
            {
                "id": skill_id,
                "name": name,
                "status": ru(raw_status),
                "result": (
                    f"Обработано: {processed}"
                    + (f"; блокировано: {blocked}" if blocked else "")
                ),
                "tone": tone,
            }
        )

    activity = [
        format_activity(row)
        for row in sorted(
            action_rows,
            key=lambda row: str(row.get("logged_at") or ""),
            reverse=True,
        )[:20]
    ]

    next_run = parse_dt(runtime_state.get("next_full_sweep_at"))
    if next_run is None:
        next_run = next_three_hour_boundary(now_moscow())

    if errors:
        health = "Требуется внимание"
    elif watcher_required and not watcher_running:
        health = "Watcher запускается"
    else:
        health = "Контур работает"

    return {
        "ok": not errors,
        "project_root": str(root),
        "updated_at": iso(now_moscow()),
        "header": {
            "health": health,
            "next_run": iso(next_run),
            "seconds_to_next": max(
                0, int((next_run - now_moscow()).total_seconds())
            ),
            "full_sweep_due": False,
            "last_status": ru(last_status_raw),
            "last_run_id": last_run.get("run_id") or "",
        },
        "kpi": {
            "new_protocols": pending_input_count(
                root, "meetings", "Source_Documents"
            ),
            "registered": len(tasks),
            "due_soon": count_timing({"Due_Soon", "Due_Today"}),
            "overdue": count_timing({"Overdue"}),
            "manual_review": count_status({"Human_Review_Required"}),
            "rework": count_status({"Rework_Required"}),
            "closed": count_status({"Closed"}),
            "notifications": sum(
                1 for row in outbox if str(row.get("status")) == "Draft"
            ),
            "results": len(results),
        },
        "stages": stage_items,
        "tasks": task_items,
        "activity": activity,
        "errors": errors,
        "system": {
            "storage_ready": not missing,
            "data_revision": data_revision(root),
            "registry_updated_at": registry_updated_at(root),
            "workspace_candidates": 1,
            "pending_commands": count_files(
                root / "input" / "commands", {".json"}
            ),
            "meeting_files": count_files(root / "input" / "meetings"),
            "result_files": count_files(root / "input" / "results"),
            "first_run": not bool(control.get("last_completed_at")),
            "auto_refresh_seconds": 10,
            "active_profile": profile,
            "profile_label": PROFILE_LABELS.get(profile, profile),
            "watcher_poll_seconds": int(
                config.get("watcher_poll_seconds") or 2
            ),
            "periodic_interval_minutes": int(
                config.get("periodic_interval_minutes") or 180
            ),
            "next_full_sweep": iso(next_run),
            "schedule_label": schedule_label(config),
            "watcher_status": watcher_status,
            "watcher_phase": heartbeat_state.get("phase") or "unknown",
            "watcher_phase_details": heartbeat_state.get("details") or "",
            "watcher_pid": heartbeat_state.get("pid"),
            "watcher_heartbeat_file": str(
                root / "archive" / ".local_runtime_heartbeat.json"
            ),
            "watcher_heartbeat_at": (
                heartbeat_state.get("heartbeat_at")
                or runtime_state.get("heartbeat_at")
            ),
            "watcher_heartbeat_age_seconds": heartbeat_age,
            "watcher_last_inject_at": runtime_state.get("last_run_at"),
            "watcher_last_error": runtime_state.get("last_error"),
            "watcher_last_run_status": runtime_state.get("last_run_status"),
            "watcher_last_run_message": runtime_state.get("last_run_message"),
            "watcher_last_command_id": runtime_state.get("last_command_id"),
            "watcher_poll_seconds": runtime_state.get(
                "poll_seconds",
                config.get("watcher_poll_seconds", 2),
            ),
            "watcher_project_root": (
                heartbeat_state.get("project_root")
                or runtime_state.get("project_root")
                or str(root)
            ),
        },
        "paths": {
            "tasks": str(registry / "03_task_register.xlsx"),
            "meetings": str(root / "input" / "meetings"),
            "results": str(root / "input" / "results"),
            "outbox": str(registry / "05_message_outbox.xlsx"),
            "logs": str(registry / "08_skill_action_log.xlsx"),
            "commands": str(root / "input" / "commands"),
        },
    }


def execute_action(action: str) -> dict[str, Any]:
    """Queue a widget command for the long-lived companion watcher.

    Route handlers run in a short-lived isolated process. The companion is
    the reliable Host Service client because Ouroboros injects its loopback
    URL and content-hash-bound skill token into the companion environment.
    """
    if action not in ACTION_NAMES:
        return {"ok": False, "message": "Неизвестное действие."}

    try:
        root = find_project_root()
    except Exception as exc:
        return {
            "ok": False,
            "message": str(exc),
            "data": get_dashboard_data(),
        }

    commands = root / "input" / "commands"
    commands.mkdir(parents=True, exist_ok=True)
    timestamp = now_moscow()
    command_id = (
        f"CMD-{timestamp:%Y%m%d-%H%M%S}-"
        f"{uuid.uuid4().hex[:6].upper()}"
    )
    command_path = commands / f"{command_id}.json"
    payload = {
        "command_id": command_id,
        "action": action,
        "action_name": ACTION_NAMES[action],
        "created_at": iso(timestamp),
        "source": SKILL_NAME,
        "status": "Queued",
        "runtime": "portable_hybrid_llm_v4.2.1",
    }

    temporary = command_path.with_suffix(".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(command_path)

    runtime_state = read_runtime_state(root)
    heartbeat_state = read_watcher_heartbeat(root)
    watcher_status = str(
        heartbeat_state.get("status")
        or runtime_state.get("status")
        or "not_started"
    )
    heartbeat = parse_dt(heartbeat_state.get("heartbeat_at"))
    heartbeat_age = (
        max(0, int((now_moscow() - heartbeat).total_seconds()))
        if heartbeat is not None
        else None
    )
    watcher_ready = (
        watcher_status == "running"
        and heartbeat_age is not None
        and heartbeat_age <= 45
    )

    return {
        "ok": True,
        "queued": True,
        "command_id": command_id,
        "run_id": None,
        "message": (
            "Команда поставлена в очередь AI watcher. "
            "Обработка языковой моделью начнётся в ближайшие несколько секунд."
            if watcher_ready
            else
            "Команда поставлена в очередь, но watcher пока не подтверждает "
            "работу. Перезапустите Ouroboros или выключите и включите skill."
        ),
        "metrics": {},
        "watcher_status": watcher_status,
        "data": get_dashboard_data(),
    }


def route_data(_request):
    return get_dashboard_data()


def route_run_full_control(_request):
    return execute_action("run_full_control")


def route_process_meetings(_request):
    return execute_action("process_meetings")


def route_check_deadlines(_request):
    return execute_action("check_deadlines")


def route_check_results(_request):
    return execute_action("check_results")


def route_refresh(_request):
    return get_dashboard_data()


def route_path(_request):
    data = get_dashboard_data()
    return {"ok": True, "paths": data.get("paths", {})}


def register(api):
    api.register_companion_process("local_runtime_watcher")

    api.register_ui_tab(
        "task_control",
        "Контроль поручений",
        icon="widgets",
        render={
            "kind": "module",
            "entry": "widget.js",
            "span": 2,
        },
    )

    api.register_route("data", handler=route_data, methods=("GET",))
    api.register_route(
        "run_full_control",
        handler=route_run_full_control,
        methods=("POST",),
    )
    api.register_route(
        "process_meetings",
        handler=route_process_meetings,
        methods=("POST",),
    )
    api.register_route(
        "check_deadlines",
        handler=route_check_deadlines,
        methods=("POST",),
    )
    api.register_route(
        "check_results",
        handler=route_check_results,
        methods=("POST",),
    )
    api.register_route("refresh", handler=route_refresh, methods=("POST",))
    api.register_route("path", handler=route_path, methods=("GET",))
