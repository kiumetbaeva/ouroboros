from __future__ import annotations

import hashlib
import importlib.util
import json
import os
import signal
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

SKILL_ROOT = Path(__file__).resolve().parents[1]
_RUNTIME_PATH = SKILL_ROOT / "runtime.py"
_RUNTIME_SPEC = importlib.util.spec_from_file_location(
    "ai_pmo_task_control_runtime_watcher_v410",
    _RUNTIME_PATH,
)
if _RUNTIME_SPEC is None or _RUNTIME_SPEC.loader is None:
    raise ModuleNotFoundError(
        f"Не удалось подготовить загрузку локального runtime: {_RUNTIME_PATH}"
    )
_RUNTIME = importlib.util.module_from_spec(_RUNTIME_SPEC)
sys.modules[_RUNTIME_SPEC.name] = _RUNTIME
_RUNTIME_SPEC.loader.exec_module(_RUNTIME)

find_project_root = _RUNTIME.find_project_root
iso = _RUNTIME.iso
load_settings = _RUNTIME.load_settings
now_moscow = _RUNTIME.now_moscow
run = _RUNTIME.run
sha256_file = _RUNTIME.sha256_file
WorkbookStore = _RUNTIME.WorkbookStore
write_runtime_state = _RUNTIME.write_runtime_state


MOSCOW = timezone(timedelta(hours=3))
STOP = False


def log(message: str) -> None:
    print(f"[AI PMO LLM watcher] {iso(now_moscow())} {message}", flush=True)


def on_stop(_signum: int, _frame: Any) -> None:
    global STOP
    STOP = True


def next_periodic_boundary(now: datetime, interval_minutes: int) -> datetime:
    local = now.astimezone(MOSCOW).replace(second=0, microsecond=0)
    minute_of_day = local.hour * 60 + local.minute
    next_value = (
        (minute_of_day // interval_minutes) + 1
    ) * interval_minutes
    day_shift, minute_in_day = divmod(next_value, 1440)
    return local.replace(hour=0, minute=0) + timedelta(
        days=day_shift,
        minutes=minute_in_day,
    )


def load_state(root: Path) -> dict[str, Any]:
    path = root / "archive" / ".local_runtime_state.json"
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data.setdefault("files", {})
            return data
    except Exception:
        pass
    return {
        "version": 4,
        "status": "starting",
        "files": {},
        "last_periodic_slot": "",
        "last_run_at": None,
        "last_error": None,
        "project_root": str(root),
    }


def stable_file(path: Path, debounce_seconds: int) -> bool:
    try:
        return time.time() - path.stat().st_mtime >= debounce_seconds
    except OSError:
        return False


def current_inputs(root: Path) -> list[tuple[str, Path]]:
    result: list[tuple[str, Path]] = []
    for kind, folder in (
        ("meeting", root / "input" / "meetings"),
        ("result", root / "input" / "results"),
    ):
        if not folder.exists():
            continue
        for path in sorted(folder.iterdir()):
            if (
                path.is_file()
                and not path.name.lower().startswith("readme")
                and path.suffix.lower() in {".txt", ".md", ".docx", ".pdf"}
            ):
                result.append((kind, path))
    return result


def processed_hashes(root: Path) -> set[str]:
    from openpyxl import load_workbook

    hashes: set[str] = set()
    source_path = root / "registry" / "09_source_register.xlsx"
    if not source_path.exists():
        return hashes

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        for sheet_name in ("Source_Documents", "Result_Inbox"):
            if sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[sheet_name]
            headers = [str(cell.value or "") for cell in ws[3]]
            try:
                hash_column = headers.index("file_hash") + 1
            except ValueError:
                continue
            for values in ws.iter_rows(
                min_row=4,
                min_col=hash_column,
                max_col=hash_column,
                values_only=True,
            ):
                value = values[0] if values else None
                if value:
                    hashes.add(str(value).lower())
    finally:
        workbook.close()
    return hashes


def changed_input_exists(
    root: Path,
    state: dict[str, Any],
    debounce_seconds: int,
) -> tuple[bool, list[str]]:
    known_processed = processed_hashes(root)
    file_state = state.setdefault("files", {})
    changed: list[str] = []

    for kind, path in current_inputs(root):
        if not stable_file(path, debounce_seconds):
            continue
        try:
            file_hash = sha256_file(path)
        except OSError:
            continue
        relative = path.relative_to(root).as_posix()
        record = file_state.setdefault(relative, {})
        record["kind"] = kind
        record["last_seen_hash"] = file_hash
        record["last_seen_at"] = iso(now_moscow())

        if file_hash.lower() in known_processed:
            record["processed"] = True
            continue

        if record.get("last_triggered_hash") != file_hash:
            changed.append(relative)

    return bool(changed), changed


def mark_triggered(
    state: dict[str, Any],
    root: Path,
    relative_paths: list[str],
) -> None:
    for relative in relative_paths:
        path = root / relative
        try:
            file_hash = sha256_file(path)
        except OSError:
            continue
        record = state.setdefault("files", {}).setdefault(relative, {})
        record["last_triggered_hash"] = file_hash
        record["last_triggered_at"] = iso(now_moscow())



def queued_commands(root: Path) -> list[Path]:
    folder = root / "input" / "commands"
    folder.mkdir(parents=True, exist_ok=True)
    return [
        path
        for path in sorted(folder.glob("*.json"))
        if path.is_file() and stable_file(path, 1)
    ]


def fail_invalid_command(root: Path, path: Path, message: str) -> None:
    failed = root / "archive" / "commands" / "failed"
    failed.mkdir(parents=True, exist_ok=True)
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, dict):
            payload = {}
    except Exception:
        payload = {}
    payload.update(
        {
            "status": "Failed",
            "completed_at": iso(now_moscow()),
            "error": message,
        }
    )
    target = failed / path.name
    target.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    try:
        path.unlink()
    except OSError:
        pass


def process_command_queue(
    root: Path,
    state: dict[str, Any],
    settings: dict[str, Any],
) -> int:
    if not bool(settings.get("widget_commands_enabled", True)):
        return 0

    processed = 0
    for command_path in queued_commands(root)[:10]:
        try:
            payload = json.loads(
                command_path.read_text(encoding="utf-8")
            )
            if not isinstance(payload, dict):
                raise ValueError("Команда должна быть JSON-объектом.")
            action = str(payload.get("action") or "").strip()
            command_id = str(
                payload.get("command_id") or command_path.stem
            )
            if action not in {
                "run_full_control",
                "process_meetings",
                "check_deadlines",
                "check_results",
            }:
                raise ValueError(f"Неизвестное действие: {action}")

            result = run(
                action,
                root=root,
                trigger_type=(
                    "widget_full"
                    if action == "run_full_control"
                    else f"widget_{action}"
                ),
                trigger_id=command_id,
                command_path=command_path,
            )
            state["last_run_at"] = iso(now_moscow())
            state["last_run_status"] = result.get("status")
            state["last_run_id"] = result.get("run_id")
            state["last_run_message"] = result.get("message")
            state["last_command_id"] = command_id
            state["last_command_action"] = action
            if not result.get("ok"):
                state["last_error"] = result.get("message")
            log(
                f"Команда {command_id}: "
                + str(result.get("message") or result)
            )
            processed += 1
        except Exception as exc:
            fail_invalid_command(root, command_path, str(exc))
            state["last_error"] = str(exc)
            log(f"Ошибка команды {command_path.name}: {exc}")
            processed += 1
    return processed


def run_watcher() -> int:
    signal.signal(signal.SIGTERM, on_stop)
    signal.signal(signal.SIGINT, on_stop)

    root = find_project_root()
    state = load_state(root)
    log(f"Рабочая папка: {root}")

    while not STOP:
        poll_seconds = 2
        try:
            store = WorkbookStore(root)
            try:
                settings = load_settings(store)
            finally:
                store.close()

            poll_seconds = max(
                1, int(settings.get("watcher_poll_seconds") or 2)
            )
            debounce_seconds = max(
                0, int(settings.get("debounce_seconds") or 15)
            )
            profile = str(
                settings.get("active_profile") or "HYBRID"
            ).upper()

            state.update(
                {
                    "status": "running",
                    "heartbeat_at": iso(now_moscow()),
                    "project_root": str(root),
                    "poll_seconds": poll_seconds,
                    "active_profile": profile,
                    "host_service_url_present": bool(
                        os.environ.get("HOST_SERVICE_URL")
                    ),
                    "host_service_token_present": bool(
                        os.environ.get("HOST_SERVICE_TOKEN")
                    ),
                    "last_error": "",
                }
            )

            process_command_queue(root, state, settings)

            files_allowed = (
                bool(settings.get("file_events_enabled", True))
                and profile in {"HYBRID", "FILES_IMMEDIATELY"}
            )
            if files_allowed:
                changed, relative_paths = changed_input_exists(
                    root,
                    state,
                    debounce_seconds,
                )
                if changed:
                    result = run(
                        "run_full_control",
                        root=root,
                        trigger_type="file_event",
                        trigger_id=hashlib.sha256(
                            "|".join(relative_paths).encode("utf-8")
                        ).hexdigest()[:12],
                    )
                    mark_triggered(state, root, relative_paths)
                    state["last_run_at"] = iso(now_moscow())
                    state["last_run_status"] = result.get("status")
                    state["last_run_id"] = result.get("run_id")
                    state["last_run_message"] = result.get("message")
                    if not result.get("ok"):
                        state["last_error"] = result.get("message")
                    log(
                        "Файловый запуск: "
                        + str(result.get("message") or result)
                    )

            periodic_allowed = (
                bool(settings.get("periodic_full_run_enabled", True))
                and profile in {"HYBRID", "EVERY_3_HOURS"}
            )
            interval = max(
                60, int(settings.get("periodic_interval_minutes") or 180)
            )
            now = now_moscow()
            slot_start_minutes = (
                (now.hour * 60 + now.minute) // interval
            ) * interval
            slot_day_shift, slot_minute = divmod(
                slot_start_minutes, 1440
            )
            slot = (
                now.replace(hour=0, minute=0, second=0, microsecond=0)
                + timedelta(days=slot_day_shift, minutes=slot_minute)
            )
            slot_key = slot.isoformat()

            minute_of_day = now.hour * 60 + now.minute
            is_periodic_boundary = minute_of_day % interval == 0

            if (
                periodic_allowed
                and is_periodic_boundary
                and state.get("last_periodic_slot") != slot_key
            ):
                result = run(
                    "run_full_control",
                    root=root,
                    trigger_type="periodic",
                    trigger_id=slot_key,
                )
                state["last_periodic_slot"] = slot_key
                state["last_run_at"] = iso(now_moscow())
                state["last_run_status"] = result.get("status")
                state["last_run_id"] = result.get("run_id")
                state["last_run_message"] = result.get("message")
                if not result.get("ok"):
                    state["last_error"] = result.get("message")
                log(
                    "Плановый запуск: "
                    + str(result.get("message") or result)
                )

            state["next_full_sweep_at"] = iso(
                next_periodic_boundary(now, interval)
            )
            write_runtime_state(root, state)
        except Exception as exc:
            state.update(
                {
                    "status": "error",
                    "heartbeat_at": iso(now_moscow()),
                    "last_error": str(exc),
                    "project_root": str(root),
                }
            )
            try:
                write_runtime_state(root, state)
            except Exception:
                pass
            log(f"Ошибка: {exc}")

        time.sleep(poll_seconds)

    state.update(
        {
            "status": "stopped",
            "heartbeat_at": iso(now_moscow()),
        }
    )
    write_runtime_state(root, state)
    log("Watcher остановлен.")
    return 0


if __name__ == "__main__":
    raise SystemExit(run_watcher())
